"""Tests for the two optional workbooks of the "Rapports a generer" section.

They land in ``{output}/excel`` so the index page picks them up like every
other export.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

import openpyxl

from src.analyzer.engine import AnalyzerReport
from src.analyzer.models import Finding, Rule
from src.core.data_dictionary_selection import DataDictionarySelection
from src.core.findings_qualification import (
    UNNAMED_ALIAS,
    FindingQualification,
    save_qualifications,
)
from src.core.models import FieldInfo, MetadataSnapshot, ObjectInfo
from src.core.orchestrator import GenerationResult, SalesforceDocumentationGenerator
from src.reporting.excel_writer import ExcelReportWriter
from src.reporting.excel_writer_findings import FINDINGS_SHEET
from src.reporting.html.renderers.index_panels import render_excel_exports


def _generator(tmp_path: Path, **overrides) -> SalesforceDocumentationGenerator:
    source = tmp_path / "source"
    source.mkdir(exist_ok=True)
    generator = SalesforceDocumentationGenerator(source, tmp_path / "out", **overrides)
    generator.alias = "MHINT"
    return generator


def _snapshot(tmp_path: Path) -> MetadataSnapshot:
    return MetadataSnapshot(
        source_dir=tmp_path / "source",
        package_roots=[],
        objects=[
            ObjectInfo(
                api_name="Account",
                label="Account",
                fields=[FieldInfo(api_name="Name", label="Name")],
            ),
            ObjectInfo(
                api_name="Lead",
                label="Lead",
                fields=[FieldInfo(api_name="Company", label="Company")],
            ),
        ],
    )


def _analyzer_report() -> AnalyzerReport:
    rule = Rule(
        id="APEX-SEC-003",
        enabled=True,
        scope="apex_class",
        category="Trusted",
        subcategory="Secure",
        severity="Critical",
        source="PMD",
        reference="",
        title="Injection SOQL",
        description="",
        rationale="",
        remediation="",
    )
    finding = Finding(
        rule=rule,
        target_kind="ApexClass",
        target_name="AccountSelector",
        message="Requete dynamique",
    )
    return AnalyzerReport(apex={"AccountSelector": [finding]})


class TestSelectedDataDictionaryExcel:
    def _run(self, tmp_path: Path, selection: DataDictionarySelection | None):
        generator = _generator(tmp_path, data_dictionary_selection=selection)
        result = GenerationResult()
        generator._generate_selected_data_dictionary_excel(
            _snapshot(tmp_path),
            ExcelReportWriter(),
            tmp_path / "out" / "excel",
            result,
        )
        return result

    def test_workbook_only_covers_the_selected_objects(self, tmp_path: Path) -> None:
        result = self._run(tmp_path, DataDictionarySelection(objects={"Account"}))

        assert len(result.selected_data_dictionary_excels) == 1
        path = result.selected_data_dictionary_excels[0]
        assert path.name == f"dataDictionnary_{date.today():%Y%m%d}.xlsx"
        assert openpyxl.load_workbook(path).sheetnames == ["Synthese", "Account"]

    def test_nothing_is_written_without_a_selection(self, tmp_path: Path) -> None:
        result = self._run(tmp_path, DataDictionarySelection())

        assert result.selected_data_dictionary_excels == []
        assert not (tmp_path / "out" / "excel").exists()

    def test_nothing_is_written_when_the_objects_are_absent_from_the_source(
        self, tmp_path: Path
    ) -> None:
        result = self._run(tmp_path, DataDictionarySelection(objects={"Ghost__c"}))

        assert result.selected_data_dictionary_excels == []
        assert not (tmp_path / "out" / "excel").exists()


class TestFindingsExcel:
    def test_workbook_is_named_after_the_alias_and_the_export_time(
        self, tmp_path: Path
    ) -> None:
        generator = _generator(tmp_path)
        result = GenerationResult()

        generator._generate_findings_excel(_analyzer_report(), result)

        assert result.findings_excel.exists()
        assert result.findings_excel.parent == tmp_path / "out" / "excel"
        assert re.fullmatch(
            r"Dewey_Findings_MHINT_\d{8}_\d{6}\.xlsx", result.findings_excel.name
        )

    def test_regenerating_does_not_overwrite_the_previous_workbook(
        self, tmp_path: Path
    ) -> None:
        """The TechLead may still be filling in the previous export, so two
        runs must produce two files."""
        generator = _generator(tmp_path)

        first = GenerationResult()
        generator._generate_findings_excel(_analyzer_report(), first)
        with patch(
            "src.reporting.excel_writer_findings.datetime"
        ) as clock:
            clock.now.return_value = datetime(2026, 8, 27, 14, 30, 5)
            second = GenerationResult()
            generator._generate_findings_excel(_analyzer_report(), second)

        assert second.findings_excel.name == "Dewey_Findings_MHINT_20260827_143005.xlsx"
        assert first.findings_excel != second.findings_excel
        assert first.findings_excel.exists()
        assert second.findings_excel.exists()

    def test_falls_back_to_a_generic_name_without_an_alias(self, tmp_path: Path) -> None:
        generator = _generator(tmp_path)
        generator.alias = ""
        result = GenerationResult()

        generator._generate_findings_excel(_analyzer_report(), result)

        assert result.findings_excel.name.startswith("Dewey_Findings_org_")

    def test_a_full_run_restores_the_imported_qualification_columns(
        self, tmp_path: Path
    ) -> None:
        """A full run overwrites the workbook, so it must re-apply the
        TechLead columns instead of handing back an empty M..S."""
        store = save_qualifications(
            tmp_path / "store.json",
            {
                "MHINT": {
                    ("AccountSelector", "APEX-SEC-003", 0): FindingQualification(
                        status="En cours", team="Squad CRM", us_number="US-1234"
                    )
                }
            },
        )
        generator = _generator(tmp_path, findings_qualifications_path=store)
        result = GenerationResult()

        generator._generate_findings_excel(_analyzer_report(), result)

        sheet = openpyxl.load_workbook(result.findings_excel)[FINDINGS_SHEET]
        assert sheet["M3"].value == "En cours"
        assert sheet["N3"].value == "Squad CRM"
        assert sheet["P3"].value == "US-1234"

    def test_a_run_without_an_alias_reads_the_unnamed_bucket(
        self, tmp_path: Path
    ) -> None:
        """The screen files alias-less runs under a shared bucket; the
        orchestrator must look them up the same way."""
        store = save_qualifications(
            tmp_path / "store.json",
            {
                UNNAMED_ALIAS: {
                    ("AccountSelector", "APEX-SEC-003", 0): FindingQualification(
                        status="Faux positif"
                    )
                }
            },
        )
        generator = _generator(tmp_path, findings_qualifications_path=store)
        generator.alias = ""
        result = GenerationResult()

        generator._generate_findings_excel(_analyzer_report(), result)

        sheet = openpyxl.load_workbook(result.findings_excel)[FINDINGS_SHEET]
        assert sheet["M3"].value == "Faux positif"

    def test_columns_stay_empty_without_a_qualification_store(
        self, tmp_path: Path
    ) -> None:
        generator = _generator(tmp_path)
        result = GenerationResult()

        generator._generate_findings_excel(_analyzer_report(), result)

        sheet = openpyxl.load_workbook(result.findings_excel)[FINDINGS_SHEET]
        assert [sheet.cell(row=3, column=col).value for col in range(13, 20)] == [
            None
        ] * 7

    def test_both_workbooks_are_listed_on_the_index_page(self, tmp_path: Path) -> None:
        generator = _generator(
            tmp_path,
            data_dictionary_selection=DataDictionarySelection(objects={"Account"}),
        )
        result = GenerationResult()
        generator._generate_selected_data_dictionary_excel(
            _snapshot(tmp_path), ExcelReportWriter(), tmp_path / "out" / "excel", result
        )
        generator._generate_findings_excel(_analyzer_report(), result)

        html = render_excel_exports(
            tmp_path / "out", tmp_path / "out" / "html" / "index.html"
        )

        assert result.selected_data_dictionary_excels[0].name in html
        assert result.findings_excel.name in html
