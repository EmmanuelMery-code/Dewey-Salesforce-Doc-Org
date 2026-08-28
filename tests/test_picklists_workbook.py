"""Tests: picklists.xlsx inventory workbook (ExcelReportWriter).

Contracts tested:
  - The sheet lists every Picklist / MultiselectPicklist field.
  - The last column repeats the picklist values as API names, in the very
    same order (and with the same separator) as the labels column.
  - A field without values keeps "-" in both value columns.
  - When fewer API names than labels are known, the label is used as the
    API name so both columns stay aligned.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.core.models import FieldInfo, ObjectInfo
from src.reporting.excel_writer import ExcelReportWriter

HEADERS = (
    "Nom de l'Objet",
    "Nom du Champ",
    "Type de Champ",
    "Picklist Globale ?",
    "Nom Picklist Globale",
    "Valeurs de la Picklist",
    "Valeurs API de la Picklist",
)


def _field(api_name: str, **kwargs: object) -> FieldInfo:
    return FieldInfo(api_name=api_name, label=api_name, data_type="Picklist", **kwargs)


def _write(objects: list[ObjectInfo], tmp_path: Path):
    output = ExcelReportWriter().write_picklists_workbook(objects, tmp_path / "picklists.xlsx")
    sheet = load_workbook(output).active
    return list(sheet.iter_rows(values_only=True))


class TestPicklistsWorkbook:
    def test_api_names_column_follows_the_labels_order(self, tmp_path: Path) -> None:
        status = _field(
            "Status__c",
            picklist_values=["Nouveau", "Ferme"],
            picklist_api_names=["New", "Closed"],
        )
        rows = _write([ObjectInfo(api_name="Account", label="Account", fields=[status])], tmp_path)

        assert rows[0] == HEADERS
        assert rows[1] == (
            "Account",
            "Status__c",
            "Picklist",
            "Non",
            "-",
            "Nouveau | Ferme",
            "New | Closed",
        )

    def test_global_value_set_field_exports_its_api_names(self, tmp_path: Path) -> None:
        priority = _field(
            "Priority__c",
            picklist_is_global=True,
            picklist_global_name="PriorityPicklist",
            picklist_values=["Haute", "Basse"],
            picklist_api_names=["HIGH", "LOW"],
        )
        rows = _write([ObjectInfo(api_name="Case", label="Case", fields=[priority])], tmp_path)

        assert rows[1] == (
            "Case",
            "Priority__c",
            "Picklist",
            "Oui",
            "PriorityPicklist",
            "Haute | Basse",
            "HIGH | LOW",
        )

    def test_multiselect_picklist_is_included(self, tmp_path: Path) -> None:
        tags = FieldInfo(
            api_name="Tags__c",
            label="Tags",
            data_type="MultiselectPicklist",
            picklist_values=["A", "B"],
            picklist_api_names=["A__c", "B__c"],
        )
        rows = _write([ObjectInfo(api_name="Lead", label="Lead", fields=[tags])], tmp_path)

        assert rows[1][2] == "MultiselectPicklist"
        assert rows[1][6] == "A__c | B__c"

    def test_field_without_values_keeps_a_dash_in_both_columns(self, tmp_path: Path) -> None:
        empty = _field("Empty__c")
        rows = _write([ObjectInfo(api_name="Account", label="Account", fields=[empty])], tmp_path)

        assert rows[1][5] == "-"
        assert rows[1][6] == "-"

    def test_missing_api_names_fall_back_to_the_labels(self, tmp_path: Path) -> None:
        partial = _field(
            "Partial__c",
            picklist_values=["Nouveau", "Ferme"],
            picklist_api_names=["New"],
        )
        rows = _write([ObjectInfo(api_name="Account", label="Account", fields=[partial])], tmp_path)

        assert rows[1][6] == "New | Ferme"

    def test_non_picklist_fields_are_skipped(self, tmp_path: Path) -> None:
        text = FieldInfo(api_name="Name", label="Name", data_type="Text")
        status = _field(
            "Status__c",
            picklist_values=["Nouveau"],
            picklist_api_names=["New"],
        )
        rows = _write(
            [ObjectInfo(api_name="Account", label="Account", fields=[text, status])], tmp_path
        )

        assert len(rows) == 2
        assert rows[1][1] == "Status__c"
