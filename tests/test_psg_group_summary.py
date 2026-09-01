"""Tests: sous-onglet "PSet Group Summary" de l'onglet Profiles & PS.

Contrats testes :
  SalesforceMetadataParser(...).parse() -> MetadataSnapshot
    les object permissions portent desormais `view_all_fields` (viewAllFields).

  build_group_access(snapshot) -> list[GroupAccess]
    les droits d'un Permission Set Group sont l'union de ceux de ses
    permission sets membres, avec la trace du permission set d'origine.

  render_security_dashboard_tab(...) -> str
    expose un sous-onglet "PSet Group Summary" contenant la matrice
    objet x groupe (CRUD + Sharing & Visibility) et l'OWD de chaque objet,
    plus une case a cocher restreignant la matrice aux objets coches dans
    l'ecran Data Dictionnary quand cette selection n'est pas vide. Le nom de
    chaque objet ouvre sa page, et son nombre de regles de partage ouvre
    l'onglet Sharing Rules filtre sur cet objet.

  ExcelReportWriter().write_psg_summary_workbook(...) -> Path
    ecrit le classeur equivalent au sous-onglet : un onglet par tableau du
    sous-onglet et de sa page de detail, liste dans les exports Excel.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.core.psg_access import COVERAGE_REASONS, SUMMARY_WORKBOOK_NAME
from src.core.utils import html_value, safe_slug
from src.parsers.salesforce_parser import SalesforceMetadataParser
from src.reporting.excel_writer import ExcelReportWriter
from src.reporting.html.renderers.index_panels import render_excel_exports
from src.reporting.html.renderers.index_tables import render_security_dashboard_tab
from src.reporting.html.renderers.psg_details import write_psg_details_page
from src.reporting.html.renderers.psg_summary import (
    DETAIL_ANCHORS,
    SHARING_RULES_PANEL_ID,
    STATUS_HELP,
    build_group_access,
)

ACCOUNT_OBJECT = """<?xml version="1.0" encoding="UTF-8"?>
<CustomObject xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Account</label>
    <sharingModel>Private</sharingModel>
</CustomObject>
"""

SALES_PERMSET = """<?xml version="1.0" encoding="UTF-8"?>
<PermissionSet xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Sales Core</label>
    <objectPermissions>
        <object>Account</object>
        <allowRead>true</allowRead>
        <allowCreate>true</allowCreate>
        <allowEdit>false</allowEdit>
        <allowDelete>false</allowDelete>
        <viewAllRecords>false</viewAllRecords>
        <modifyAllRecords>false</modifyAllRecords>
        <viewAllFields>false</viewAllFields>
    </objectPermissions>
</PermissionSet>
"""

ADMIN_PERMSET = """<?xml version="1.0" encoding="UTF-8"?>
<PermissionSet xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Sales Admin</label>
    <objectPermissions>
        <object>Account</object>
        <allowRead>true</allowRead>
        <allowCreate>false</allowCreate>
        <allowEdit>true</allowEdit>
        <allowDelete>true</allowDelete>
        <viewAllRecords>true</viewAllRecords>
        <modifyAllRecords>true</modifyAllRecords>
        <viewAllFields>true</viewAllFields>
    </objectPermissions>
</PermissionSet>
"""

ACCOUNT_SHARING_RULES = """<?xml version="1.0" encoding="UTF-8"?>
<SharingRules xmlns="http://soap.sforce.com/2006/04/metadata">
    <sharingCriteriaRules>
        <fullName>Account_Partners</fullName>
        <label>Account Partners</label>
        <description>Partage des comptes partenaires</description>
    </sharingCriteriaRules>
</SharingRules>
"""

SALES_GROUP = """<?xml version="1.0" encoding="UTF-8"?>
<PermissionSetGroup xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Sales Team</label>
    <status>Updated</status>
    <permissionSets>Sales_Core</permissionSets>
    <permissionSets>Sales_Admin</permissionSets>
    <permissionSets>Missing_Set</permissionSets>
</PermissionSetGroup>
"""


def _build_source(tmp_path: Path) -> Path:
    source = tmp_path / "source"
    (source / "objects" / "Account").mkdir(parents=True, exist_ok=True)
    (source / "permissionsets").mkdir(parents=True, exist_ok=True)
    (source / "permissionsetgroups").mkdir(parents=True, exist_ok=True)
    (source / "objects" / "Account" / "Account.object-meta.xml").write_text(
        ACCOUNT_OBJECT, encoding="utf-8"
    )
    (source / "permissionsets" / "Sales_Core.permissionset-meta.xml").write_text(
        SALES_PERMSET, encoding="utf-8"
    )
    (source / "permissionsets" / "Sales_Admin.permissionset-meta.xml").write_text(
        ADMIN_PERMSET, encoding="utf-8"
    )
    (source / "permissionsetgroups" / "Sales_Team.permissionsetgroup-meta.xml").write_text(
        SALES_GROUP, encoding="utf-8"
    )
    return source


def _snapshot(tmp_path: Path):
    return SalesforceMetadataParser(_build_source(tmp_path)).parse()


class TestObjectPermissionParsing:
    def test_view_all_fields_is_parsed(self, tmp_path: Path) -> None:
        snapshot = _snapshot(tmp_path)

        by_name = {item.name: item for item in snapshot.permission_sets}
        admin_perm = by_name["Sales_Admin"].object_permissions[0]
        core_perm = by_name["Sales_Core"].object_permissions[0]

        assert admin_perm.view_all_fields is True
        assert core_perm.view_all_fields is False


class TestGroupAccessAggregation:
    def test_group_permissions_are_the_union_of_its_permission_sets(self, tmp_path: Path) -> None:
        accesses = build_group_access(_snapshot(tmp_path))

        assert [item.group.name for item in accesses] == ["Sales_Team"]
        account = accesses[0].objects["Account"]
        assert account.granted("allow_create") is True
        assert account.granted("allow_delete") is True
        assert account.granted("view_all_records") is True
        assert account.granted("modify_all_records") is True
        assert account.granted("view_all_fields") is True

    def test_each_permission_is_traced_back_to_its_permission_set(self, tmp_path: Path) -> None:
        accesses = build_group_access(_snapshot(tmp_path))
        account = accesses[0].objects["Account"]

        assert account.sources("allow_create") == ["Sales_Core"]
        assert account.sources("modify_all_records") == ["Sales_Admin"]
        assert account.contributors == ["Sales_Admin", "Sales_Core"]

    def test_permission_sets_absent_from_the_analysis_are_flagged(self, tmp_path: Path) -> None:
        accesses = build_group_access(_snapshot(tmp_path))

        assert accesses[0].unresolved_permission_sets == ["Missing_Set"]
        assert accesses[0].resolved_permission_sets == ["Sales_Core", "Sales_Admin"]


class TestGroupSummarySubTab:
    def _render(self, tmp_path: Path) -> str:
        return render_security_dashboard_tab(
            _snapshot(tmp_path),
            tmp_path / "html" / "index.html",
            {"psg_list": tmp_path / "html" / "psg_list.html"},
            None,
        )

    def test_sub_tab_is_added_to_the_profiles_and_ps_tab(self, tmp_path: Path) -> None:
        panel = self._render(tmp_path).replace("'", '"')

        assert f'id="index-security-panel-{safe_slug("PSet Group Summary")}"' in panel
        assert ">PSet Group Summary</button>" in panel

    def test_matrix_shows_crud_and_sharing_visibility_per_object_and_group(
        self, tmp_path: Path
    ) -> None:
        panel = self._render(tmp_path)

        assert "Sharing &amp; Visibility" in panel
        assert "Sales Team" in panel
        assert "View All / Voir tout" in panel
        assert "Modify All / Modifier tout" in panel
        assert "View All Fields / Voir tous les champs" in panel

    def test_matrix_includes_the_object_wide_default_sharing_model(self, tmp_path: Path) -> None:
        panel = self._render(tmp_path)

        assert "Private" in panel

    def test_objects_without_any_group_right_are_still_listed(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        (source / "objects" / "Lead").mkdir(parents=True, exist_ok=True)
        (source / "objects" / "Lead" / "Lead.object-meta.xml").write_text(
            ACCOUNT_OBJECT.replace("Account", "Lead"), encoding="utf-8"
        )
        snapshot = SalesforceMetadataParser(source).parse()

        panel = render_security_dashboard_tab(snapshot, tmp_path / "index.html", None, None)

        assert "data-object='lead'" in panel
        assert "data-covered='0'" in panel
        assert "<input type='checkbox' id='psg-matrix-covered-only'>" in panel
        assert "sur 2 objet(s) listes" in panel

    def test_data_dictionary_selection_adds_the_filter_checkbox(self, tmp_path: Path) -> None:
        panel = render_security_dashboard_tab(
            _snapshot(tmp_path),
            tmp_path / "index.html",
            None,
            None,
            {"Account"},
        )

        assert "<input type='checkbox' id='psg-matrix-selected-only'>" in panel
        assert "Filtrer sur les objets selectionnes" in panel
        assert "data-selected='1'" in panel

    def test_no_checkbox_without_data_dictionary_selection(self, tmp_path: Path) -> None:
        panel = self._render(tmp_path)

        assert "<input type='checkbox' id='psg-matrix-selected-only'>" not in panel
        assert "Filtrer sur les objets selectionnes" not in panel
        assert "data-selected='0'" in panel

    def test_group_summary_is_empty_when_no_group_is_analysed(self, tmp_path: Path) -> None:
        source = tmp_path / "empty"
        source.mkdir()
        snapshot = SalesforceMetadataParser(source).parse()

        panel = render_security_dashboard_tab(snapshot, tmp_path / "index.html", None, None)

        assert "Aucun Permission Set Group analyse" in panel


class TestMatrixLinks:
    def _render(self, tmp_path: Path, object_pages: dict[str, Path] | None = None) -> str:
        return render_security_dashboard_tab(
            _snapshot(tmp_path),
            tmp_path / "html" / "index.html",
            None,
            None,
            None,
            None,
            object_pages,
        )

    def test_object_name_links_to_its_page(self, tmp_path: Path) -> None:
        panel = self._render(
            tmp_path, {"Account": tmp_path / "html" / "objects" / "Account.html"}
        )

        assert "<a href='objects/Account.html'" in panel

    def test_object_name_stays_plain_text_without_a_page(self, tmp_path: Path) -> None:
        panel = self._render(tmp_path)

        assert "<th scope='row' class='psg-sticky'>Account</th>" in panel

    def test_sharing_rule_count_links_to_the_filtered_sharing_rules_tab(
        self, tmp_path: Path
    ) -> None:
        source = _build_source(tmp_path)
        (source / "sharingRules").mkdir(parents=True, exist_ok=True)
        (source / "sharingRules" / "Account.sharingRules-meta.xml").write_text(
            ACCOUNT_SHARING_RULES, encoding="utf-8"
        )
        snapshot = SalesforceMetadataParser(source).parse()

        panel = render_security_dashboard_tab(
            snapshot, tmp_path / "html" / "index.html", None, None
        )

        assert f"<a href='#{SHARING_RULES_PANEL_ID}'" in panel
        assert "data-psg-sharing-filter='Account'" in panel
        assert "search.value = link.dataset.psgSharingFilter" in panel

    def test_no_link_when_the_object_has_no_sharing_rule(self, tmp_path: Path) -> None:
        panel = self._render(tmp_path)

        assert "data-psg-sharing-filter=" not in panel


class TestStatusLegend:
    def test_legend_lists_every_possible_group_status(self, tmp_path: Path) -> None:
        panel = render_security_dashboard_tab(
            _snapshot(tmp_path), tmp_path / "index.html", None, None
        )

        assert "<em>Statut</em> du groupe" in panel
        for value, description in STATUS_HELP:
            assert value in panel
            # The description lives in a title attribute, hence the escaping.
            assert html_value(description) in panel


class TestDetailsPage:
    def _write(self, tmp_path: Path) -> str:
        output_dir = tmp_path / "html"
        output_dir.mkdir(parents=True, exist_ok=True)
        page = write_psg_details_page(
            _snapshot(tmp_path), output_dir, output_dir / "assets"
        )
        return page.read_text(encoding="utf-8")

    def test_every_kpi_title_links_to_a_section_of_the_details_page(
        self, tmp_path: Path
    ) -> None:
        details = tmp_path / "html" / "psg_summary_details.html"
        panel = render_security_dashboard_tab(
            _snapshot(tmp_path),
            tmp_path / "html" / "index.html",
            {"psg_details": details},
            None,
        )
        content = self._write(tmp_path)

        for anchor in DETAIL_ANCHORS.values():
            assert f"psg_summary_details.html#{anchor}" in panel
            assert f"id='{anchor}'" in content

    def test_details_page_explains_the_object_coverage(self, tmp_path: Path) -> None:
        content = self._write(tmp_path)

        assert "Pourquoi un objet est-il couvert ou non ?" in content
        assert "Objets analyses sans aucun droit via un groupe" in content
        assert "Account" in content

    def test_details_page_documents_every_status(self, tmp_path: Path) -> None:
        content = self._write(tmp_path)

        for value, _description in STATUS_HELP:
            assert value in content

    def test_details_page_lists_unresolved_permission_sets(self, tmp_path: Path) -> None:
        content = self._write(tmp_path)

        assert "Missing_Set" in content
        assert "Non analyse" in content

    def test_back_link_targets_the_group_summary_sub_tab(self, tmp_path: Path) -> None:
        content = self._write(tmp_path)

        anchor = f"index-security-panel-{safe_slug('PSet Group Summary')}"
        assert f"#{anchor}" in content


class TestPsgSummaryWorkbook:
    def _workbook(self, tmp_path: Path, selected: set[str] | None = None):
        path = ExcelReportWriter().write_psg_summary_workbook(
            _snapshot(tmp_path),
            tmp_path / "excel" / SUMMARY_WORKBOOK_NAME,
            selected_objects=selected,
        )
        return load_workbook(path)

    @staticmethod
    def _rows(sheet) -> list[list[object]]:
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def test_one_sheet_per_table_of_the_summary_and_details_pages(
        self, tmp_path: Path
    ) -> None:
        workbook = self._workbook(tmp_path)

        assert workbook.sheetnames == [
            "Synthese",
            "Matrice",
            "DroitsParGroupe",
            "Groupes",
            "PermissionSets",
            "Couverture",
            "ModifyAll",
            "ViewAll",
            "Legende",
        ]

    def test_matrix_sheet_holds_one_column_pair_per_group(self, tmp_path: Path) -> None:
        rows = self._rows(self._workbook(tmp_path)["Matrice"])

        assert rows[0] == [
            "Objet",
            "OWD",
            "Regles de partage",
            "Couvert",
            "Sales Team - CRUD",
            "Sales Team - Sharing & Visibility",
        ]
        assert rows[1] == ["Account", "Private", 0, "Oui", "C R U D", "VA MA VAF"]

    def test_kpi_sheet_mirrors_the_cards_of_the_sub_tab(self, tmp_path: Path) -> None:
        rows = self._rows(self._workbook(tmp_path)["Synthese"])
        by_label = {row[0]: row for row in rows[1:]}

        assert by_label["Permission Set Groups"][1] == 1
        assert by_label["Permission Sets membres"][1] == 3
        assert by_label["Permission Sets membres"][2] == "1 non analyse(s)"
        assert by_label["Objets couverts"][1] == 1
        assert by_label["Couples groupe/objet avec Modify All"][1] == 1
        # Chaque cadran renvoie vers l'onglet qui detaille son chiffre.
        assert by_label["Couples groupe/objet avec View All"][3] == "ViewAll"

    def test_wide_access_sheets_name_the_granting_permission_set(
        self, tmp_path: Path
    ) -> None:
        workbook = self._workbook(tmp_path)

        assert self._rows(workbook["ModifyAll"])[1] == [
            "Sales Team",
            "Updated",
            "Account",
            "Sales_Admin",
        ]
        assert self._rows(workbook["ViewAll"])[1][2] == "Account"

    def test_permission_sets_sheet_flags_the_unanalysed_member(self, tmp_path: Path) -> None:
        rows = self._rows(self._workbook(tmp_path)["PermissionSets"])
        by_name = {row[0]: row for row in rows[1:]}

        assert by_name["Missing_Set"][1] == "Non analyse"
        assert by_name["Sales_Admin"][1] == "Analyse"

    def test_objects_without_any_group_right_are_listed_as_uncovered(
        self, tmp_path: Path
    ) -> None:
        source = _build_source(tmp_path)
        (source / "objects" / "Lead").mkdir(parents=True, exist_ok=True)
        (source / "objects" / "Lead" / "Lead.object-meta.xml").write_text(
            ACCOUNT_OBJECT.replace("Account", "Lead"), encoding="utf-8"
        )
        path = ExcelReportWriter().write_psg_summary_workbook(
            SalesforceMetadataParser(source).parse(),
            tmp_path / "excel" / SUMMARY_WORKBOOK_NAME,
        )
        rows = self._rows(load_workbook(path)["Couverture"])
        by_object = {row[0]: row for row in rows[1:]}

        assert by_object["Lead"][3] is None
        assert by_object["Account"][3] == "Oui"

    def test_data_dictionary_selection_becomes_a_filterable_column(
        self, tmp_path: Path
    ) -> None:
        rows = self._rows(self._workbook(tmp_path, selected={"Account"})["Matrice"])

        assert "Selectionne (Data Dictionary)" in rows[0]
        assert rows[1][rows[0].index("Selectionne (Data Dictionary)")] == "Oui"

    def test_legend_sheet_explains_statuses_and_object_coverage(
        self, tmp_path: Path
    ) -> None:
        rows = self._rows(self._workbook(tmp_path)["Legende"])
        values = {row[1]: row[2] for row in rows[1:]}

        for status, description in STATUS_HELP:
            assert values[status] == description
        for title, explanation in COVERAGE_REASONS:
            assert values[title] == explanation

    def test_workbook_is_listed_in_the_excel_exports_section(self, tmp_path: Path) -> None:
        ExcelReportWriter().write_psg_summary_workbook(
            _snapshot(tmp_path), tmp_path / "excel" / SUMMARY_WORKBOOK_NAME
        )

        listing = render_excel_exports(tmp_path, tmp_path / "html" / "index.html")

        assert SUMMARY_WORKBOOK_NAME in listing

    def test_sub_tab_links_to_the_workbook_once_generated(self, tmp_path: Path) -> None:
        workbook = tmp_path / "excel" / SUMMARY_WORKBOOK_NAME
        ExcelReportWriter().write_psg_summary_workbook(_snapshot(tmp_path), workbook)

        panel = render_security_dashboard_tab(
            _snapshot(tmp_path),
            tmp_path / "html" / "index.html",
            None,
            None,
            None,
            workbook,
        )

        assert "Telecharger le classeur Excel equivalent" in panel
        assert SUMMARY_WORKBOOK_NAME in panel

    def test_no_workbook_link_when_the_workbook_is_missing(self, tmp_path: Path) -> None:
        panel = render_security_dashboard_tab(
            _snapshot(tmp_path),
            tmp_path / "html" / "index.html",
            None,
            None,
            None,
            tmp_path / "excel" / SUMMARY_WORKBOOK_NAME,
        )

        assert "Telecharger le classeur Excel equivalent" not in panel
