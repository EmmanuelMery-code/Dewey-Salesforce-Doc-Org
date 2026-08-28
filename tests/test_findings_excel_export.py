"""Tests for the findings qualification workbook and its on-disk cache."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from src.analyzer.engine import AnalyzerReport
from src.analyzer.models import Finding, Rule
from src.core.findings_cache import load_findings_cache, save_findings_cache
from src.reporting.excel_writer_findings import (
    FINDINGS_SHEET,
    LEGEND_SHEET,
    FindingsExcelWriter,
)


def _rule(rule_id: str = "APEX-SEC-003", severity: str = "Critical", **overrides) -> Rule:
    defaults = dict(
        id=rule_id,
        enabled=True,
        scope="apex_class",
        category="Trusted",
        subcategory="Secure",
        severity=severity,
        source="PMD ApexSOQLInjection",
        reference="https://example.com/rule",
        title="Risque d'injection SOQL detecte",
        description="Requete dynamique sans echappement.",
        rationale="Un attaquant peut manipuler la requete.",
        remediation="Utiliser des variables de liaison.",
    )
    defaults.update(overrides)
    return Rule(**defaults)


def _finding(rule: Rule, target_name: str, **overrides) -> Finding:
    defaults = dict(
        rule=rule,
        target_kind="ApexClass",
        target_name=target_name,
        message="Risque d'injection SOQL dans une requete dynamique.",
        details=["Database.query() sans echappement."],
    )
    defaults.update(overrides)
    return Finding(**defaults)


class TestFindingsWorkbook:
    def test_workbook_has_the_two_expected_sheets(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "AccountSelector")], tmp_path / "findings.xlsx"
        )

        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames == [FINDINGS_SHEET, LEGEND_SHEET]

    def test_header_carries_the_alias_and_the_run_date(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "AccountSelector")],
            tmp_path / "findings.xlsx",
            alias="MHINT",
            run_date=date(2026, 7, 10),
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert sheet["A1"].value == "Finding (MHINT 10/07/2026)"
        assert sheet["M1"].value == "Qualification"
        assert sheet["Q1"].value == "US"
        assert sheet.freeze_panes == "A3"

    def test_finding_row_maps_the_rule_onto_the_first_columns(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "AccountSelector")], tmp_path / "findings.xlsx"
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert sheet["A3"].value == "Critique"
        assert sheet["B3"].value == "Trusted - Secure"
        assert sheet["C3"].value == "APEX-SEC-003"
        assert sheet["D3"].value == "ApexClass"
        assert sheet["E3"].value == "AccountSelector"
        assert sheet["F3"].value == "Risque d'injection SOQL detecte"
        assert sheet["G3"].value == "Risque d'injection SOQL dans une requete dynamique."
        assert sheet["H3"].value == "Un attaquant peut manipuler la requete."
        assert sheet["I3"].value == "Utiliser des variables de liaison."
        assert sheet["J3"].value == "PMD ApexSOQLInjection"
        assert sheet["K3"].value == "https://example.com/rule"
        assert sheet["L3"].value == "Database.query() sans echappement."

    def test_qualification_and_us_columns_are_left_empty(self, tmp_path: Path) -> None:
        findings = [_finding(_rule(), "AccountSelector"), _finding(_rule(), "Beta")]
        path = FindingsExcelWriter().write_findings_workbook(
            findings, tmp_path / "findings.xlsx"
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        for row in range(3, 3 + len(findings)):
            assert [sheet.cell(row=row, column=col).value for col in range(13, 20)] == [
                None
            ] * 7

    def test_rows_are_sorted_by_severity_then_component(self, tmp_path: Path) -> None:
        findings = [
            _finding(_rule("OBJ-DOC-001", "Minor"), "Zebra__c", target_kind="Object"),
            _finding(_rule("APEX-REL-003", "Major"), "Beta"),
            _finding(_rule("APEX-REL-002", "Major"), "Alpha"),
            _finding(_rule(), "AccountSelector"),
        ]
        path = FindingsExcelWriter().write_findings_workbook(
            findings, tmp_path / "findings.xlsx"
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert [sheet.cell(row=row, column=5).value for row in range(3, 7)] == [
            "AccountSelector",
            "Alpha",
            "Beta",
            "Zebra__c",
        ]

    def test_status_column_offers_the_four_statuses(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "AccountSelector")], tmp_path / "findings.xlsx"
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        validation = sheet.data_validations.dataValidation[0]
        assert validation.formula1 == '"À traiter,Faux positif,En cours,Terminé"'
        assert str(validation.sqref) == "M3"

    def test_legend_sheet_is_static(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "AccountSelector")], tmp_path / "findings.xlsx"
        )

        sheet = openpyxl.load_workbook(path)[LEGEND_SHEET]
        assert sheet["A2"].value == "STATUTS"
        assert sheet["A8"].value == "WORKFLOW"
        assert sheet["A3"].value == "À traiter"
        assert sheet["A10"].value == "Vrai sujet"


class TestFindingsCache:
    def test_round_trip_preserves_findings_and_context(self, tmp_path: Path) -> None:
        finding = _finding(_rule(), "AccountSelector", source_path=tmp_path / "Foo.cls", line=42)
        report = AnalyzerReport(apex={"AccountSelector": [finding]})

        cache_path = save_findings_cache(report, tmp_path / "cache.json", alias="MHINT")
        cached = load_findings_cache(cache_path)

        assert cached is not None
        assert cached.alias == "MHINT"
        assert cached.generated_at == date.today()
        assert len(cached.findings) == 1
        restored = cached.findings[0]
        assert restored.rule.id == "APEX-SEC-003"
        assert restored.rule.remediation == "Utiliser des variables de liaison."
        assert restored.target_name == "AccountSelector"
        assert restored.details == ["Database.query() sans echappement."]
        assert restored.line == 42

    def test_missing_or_empty_cache_returns_none(self, tmp_path: Path) -> None:
        assert load_findings_cache(tmp_path / "absent.json") is None

        empty = save_findings_cache(AnalyzerReport(), tmp_path / "empty.json")
        assert load_findings_cache(empty) is None

    def test_no_report_writes_nothing(self, tmp_path: Path) -> None:
        target = tmp_path / "cache.json"
        assert save_findings_cache(None, target) is None
        assert not target.exists()
