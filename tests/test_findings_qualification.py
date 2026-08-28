"""Tests for the findings qualification round trip (export -> Excel -> import).

Covers the matching key (component + rule, disambiguated by row order), the
per-org store, the per-alias findings caches and the fact that an imported
qualification is written back on the next export.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from src.analyzer.engine import AnalyzerReport
from src.analyzer.models import Finding, Rule
from src.core.findings_cache import (
    CACHE_FILENAME,
    findings_cache_path,
    load_all_findings_caches,
    save_findings_cache,
)
from src.core.findings_qualification import (
    FindingQualification,
    assign_keys,
    finding_keys,
    load_qualifications,
    save_qualifications,
)
from src.reporting.excel_reader_findings import (
    FindingsWorkbookError,
    read_findings_qualifications,
)
from src.reporting.excel_writer_findings import FINDINGS_SHEET, FindingsExcelWriter


def _rule(rule_id: str = "APEX-SEC-003", severity: str = "Critical") -> Rule:
    return Rule(
        id=rule_id,
        enabled=True,
        scope="apex_class",
        category="Trusted",
        subcategory="Secure",
        severity=severity,
        source="PMD ApexSOQLInjection",
        reference="https://example.com/rule",
        title="Risque d'injection SOQL detecte",
        description="Requete dynamique sans echappement.",
        rationale="Un attaquant peut manipuler la requete.",
        remediation="Utiliser des variables de liaison.",
    )


def _finding(rule: Rule, target_name: str, message: str = "Occurrence") -> Finding:
    return Finding(
        rule=rule,
        target_kind="ApexClass",
        target_name=target_name,
        message=message,
        details=["Database.query() sans echappement."],
    )


def _qualification(status: str = "En cours", **overrides) -> FindingQualification:
    defaults = dict(
        status=status,
        team="Squad CRM",
        target_sprint="S26-04",
        us_number="US-1234",
        us_title="Securiser AccountSelector",
        us_description="Remplacer la requete dynamique.",
        acceptance_criteria="Plus aucune Database.query() non echappee.",
    )
    defaults.update(overrides)
    return FindingQualification(**defaults)


class TestMatchingKey:
    def test_distinct_pairs_all_start_at_occurrence_zero(self) -> None:
        keys = assign_keys([("Alpha", "R1"), ("Beta", "R1"), ("Alpha", "R2")])

        assert keys == [("Alpha", "R1", 0), ("Beta", "R1", 0), ("Alpha", "R2", 0)]

    def test_duplicate_pairs_are_disambiguated_by_row_order(self) -> None:
        keys = assign_keys(
            [("Alpha", "R1"), ("Alpha", "R1"), ("Beta", "R1"), ("Alpha", "R1")]
        )

        assert keys == [
            ("Alpha", "R1", 0),
            ("Alpha", "R1", 1),
            ("Beta", "R1", 0),
            ("Alpha", "R1", 2),
        ]

    def test_keys_are_trimmed_so_excel_padding_still_matches(self) -> None:
        assert assign_keys([("  Alpha ", " R1 ")]) == [("Alpha", "R1", 0)]

    def test_finding_keys_follow_the_given_order(self) -> None:
        rule = _rule()
        findings = [_finding(rule, "Alpha"), _finding(rule, "Alpha")]

        assert finding_keys(findings) == [
            ("Alpha", "APEX-SEC-003", 0),
            ("Alpha", "APEX-SEC-003", 1),
        ]


class TestStore:
    def test_round_trip_keeps_every_column_per_alias(self, tmp_path: Path) -> None:
        store = {
            "MHINT": {("Alpha", "R1", 0): _qualification()},
            "AG2R": {("Beta", "R2", 1): _qualification(status="Terminé")},
        }

        path = save_qualifications(tmp_path / "store.json", store)
        reloaded = load_qualifications(path)

        assert set(reloaded) == {"MHINT", "AG2R"}
        assert reloaded["MHINT"][("Alpha", "R1", 0)] == _qualification()
        assert reloaded["AG2R"][("Beta", "R2", 1)].status == "Terminé"

    def test_empty_qualifications_are_not_persisted(self, tmp_path: Path) -> None:
        store = {"MHINT": {("Alpha", "R1", 0): FindingQualification()}}

        reloaded = load_qualifications(save_qualifications(tmp_path / "s.json", store))

        assert reloaded == {}

    def test_unreadable_store_yields_an_empty_store(self, tmp_path: Path) -> None:
        broken = tmp_path / "broken.json"
        broken.write_text("{not json", encoding="utf-8")

        assert load_qualifications(broken) == {}
        assert load_qualifications(tmp_path / "absent.json") == {}


class TestWorkbookRoundTrip:
    def test_exported_qualifications_are_read_back_identically(
        self, tmp_path: Path
    ) -> None:
        rule = _rule()
        findings = [_finding(rule, "Alpha"), _finding(rule, "Beta")]
        qualifications = {
            ("Alpha", "APEX-SEC-003", 0): _qualification(),
            ("Beta", "APEX-SEC-003", 0): _qualification(status="Faux positif"),
        }

        path = FindingsExcelWriter().write_findings_workbook(
            findings, tmp_path / "f.xlsx", qualifications=qualifications
        )

        assert read_findings_qualifications(path) == qualifications

    def test_duplicated_component_and_rule_keep_their_own_values(
        self, tmp_path: Path
    ) -> None:
        """The (component, rule) pair repeats, so only the row order tells the
        two findings apart — on export and on import alike."""
        rule = _rule()
        findings = [
            _finding(rule, "Alpha", "premiere occurrence"),
            _finding(rule, "Alpha", "seconde occurrence"),
        ]
        qualifications = {
            ("Alpha", "APEX-SEC-003", 0): _qualification(us_number="US-1"),
            ("Alpha", "APEX-SEC-003", 1): _qualification(us_number="US-2"),
        }

        path = FindingsExcelWriter().write_findings_workbook(
            findings, tmp_path / "f.xlsx", qualifications=qualifications
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert [sheet.cell(row=row, column=16).value for row in (3, 4)] == [
            "US-1",
            "US-2",
        ]
        assert read_findings_qualifications(path) == qualifications

    def test_qualifications_land_in_columns_m_to_s(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "Alpha")],
            tmp_path / "f.xlsx",
            qualifications={("Alpha", "APEX-SEC-003", 0): _qualification()},
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert [sheet.cell(row=3, column=col).value for col in range(13, 20)] == [
            "En cours",
            "Squad CRM",
            "S26-04",
            "US-1234",
            "Securiser AccountSelector",
            "Remplacer la requete dynamique.",
            "Plus aucune Database.query() non echappee.",
        ]

    def test_findings_without_a_stored_qualification_stay_empty(
        self, tmp_path: Path
    ) -> None:
        rule = _rule()
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(rule, "Alpha"), _finding(rule, "Beta")],
            tmp_path / "f.xlsx",
            qualifications={("Alpha", "APEX-SEC-003", 0): _qualification()},
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert sheet["M3"].value == "En cours"
        assert [sheet.cell(row=4, column=col).value for col in range(13, 20)] == [
            None
        ] * 7

    def test_untouched_export_imports_nothing(self, tmp_path: Path) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding(_rule(), "Alpha")], tmp_path / "f.xlsx"
        )

        assert read_findings_qualifications(path) == {}

    def test_a_workbook_without_the_findings_sheet_is_rejected(
        self, tmp_path: Path
    ) -> None:
        other = tmp_path / "other.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "Autre"
        workbook.save(other)

        with pytest.raises(FindingsWorkbookError):
            read_findings_qualifications(other)

    def test_a_non_excel_file_is_rejected(self, tmp_path: Path) -> None:
        text = tmp_path / "notes.txt"
        text.write_text("pas un classeur", encoding="utf-8")

        with pytest.raises(FindingsWorkbookError):
            read_findings_qualifications(text)


class TestPerAliasCaches:
    def test_each_alias_gets_its_own_cache(self, tmp_path: Path) -> None:
        for alias, target in (("MHINT", "Alpha"), ("AG2R", "Beta")):
            save_findings_cache(
                AnalyzerReport(apex={target: [_finding(_rule(), target)]}),
                findings_cache_path(tmp_path, alias),
                alias=alias,
            )

        caches = load_all_findings_caches(tmp_path)

        assert set(caches) == {"MHINT", "AG2R"}
        assert caches["MHINT"].findings[0].target_name == "Alpha"
        assert caches["AG2R"].findings[0].target_name == "Beta"

    def test_aliases_unusable_as_filenames_still_get_a_cache(
        self, tmp_path: Path
    ) -> None:
        path = findings_cache_path(tmp_path, "prod/eu — 1")
        save_findings_cache(
            AnalyzerReport(apex={"Alpha": [_finding(_rule(), "Alpha")]}),
            path,
            alias="prod/eu — 1",
        )

        assert path.name == "prod_eu_1.json"
        assert load_all_findings_caches(tmp_path)["prod/eu — 1"] is not None

    def test_the_legacy_single_file_cache_is_still_read(self, tmp_path: Path) -> None:
        save_findings_cache(
            AnalyzerReport(apex={"Alpha": [_finding(_rule(), "Alpha")]}),
            tmp_path / CACHE_FILENAME,
            alias="Legacy",
        )

        assert "Legacy" in load_all_findings_caches(tmp_path)

    def test_a_per_alias_cache_wins_over_the_legacy_one(self, tmp_path: Path) -> None:
        save_findings_cache(
            AnalyzerReport(apex={"Old": [_finding(_rule(), "Old")]}),
            tmp_path / CACHE_FILENAME,
            alias="MHINT",
        )
        save_findings_cache(
            AnalyzerReport(apex={"New": [_finding(_rule(), "New")]}),
            findings_cache_path(tmp_path, "MHINT"),
            alias="MHINT",
        )

        caches = load_all_findings_caches(tmp_path)

        assert caches["MHINT"].findings[0].target_name == "New"

    def test_no_cache_at_all_yields_nothing(self, tmp_path: Path) -> None:
        assert load_all_findings_caches(tmp_path) == {}


def test_qualification_survives_a_full_export_import_export_cycle(
    tmp_path: Path,
) -> None:
    """The scenario the screen drives: export, qualify in Excel, import,
    re-export and find the qualification back in place."""
    rule = _rule()
    findings = [_finding(rule, "Alpha"), _finding(rule, "Alpha")]
    writer = FindingsExcelWriter()

    first = writer.write_findings_workbook(
        findings, tmp_path / "first.xlsx", alias="MHINT", run_date=date(2026, 7, 10)
    )

    # The TechLead fills the second occurrence only.
    workbook = openpyxl.load_workbook(first)
    sheet = workbook[FINDINGS_SHEET]
    sheet["M4"] = "Faux positif"
    sheet["N4"] = "Squad CRM"
    workbook.save(first)

    imported = read_findings_qualifications(first)
    assert list(imported) == [("Alpha", "APEX-SEC-003", 1)]

    store_path = save_qualifications(tmp_path / "store.json", {"MHINT": imported})
    stored = load_qualifications(store_path)["MHINT"]

    second = writer.write_findings_workbook(
        findings, tmp_path / "second.xlsx", alias="MHINT", qualifications=stored
    )

    sheet = openpyxl.load_workbook(second)[FINDINGS_SHEET]
    assert sheet["M3"].value is None
    assert sheet["M4"].value == "Faux positif"
    assert sheet["N4"].value == "Squad CRM"
