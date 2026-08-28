"""Tests for the "Utilisé dans une automatisation ?" column.

Contract tested:
  src.core.field_automation_usage
    field_automation_usages(dependencies) -> {"object.field": [labels]}
      the automation/code kinds of the impact analysis that reference each
      field, deduplicated and in a stable display order. Presentation kinds
      (Layout, FlexiPage, Report) are ignored.
    assign_field_automation_usages(snapshot)
      stores those labels on every FieldInfo of the snapshot.

  The per-object fields sheet of the Data Dictionary workbook renders them in
  a column sitting between "Description" and "Commentaire Dewey".
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.core.field_automation_usage import (
    assign_field_automation_usages,
    field_automation_usages,
)
from src.core.models import Dependency, FieldInfo, MetadataSnapshot, ObjectInfo
from src.parsers.salesforce_parser import SalesforceMetadataParser
from src.reporting.excel_writer import ExcelReportWriter


def _dependency(source_kind: str, target: str, source_name: str = "Src") -> Dependency:
    return Dependency(
        source_name=source_name,
        source_kind=source_kind,
        target_name=target,
        target_kind="Field",
    )


class TestFieldAutomationUsages:
    def test_labels_are_deduplicated_and_ordered(self) -> None:
        """A Flow referencing the same field a dozen times is still one label,
        and the order does not follow the scan order."""
        usages = field_automation_usages(
            [
                _dependency("class", "Compte__c.Statut__c", "ServiceA"),
                _dependency("Flow", "Compte__c.Statut__c", "MonFlow"),
                _dependency("Flow", "Compte__c.Statut__c", "MonFlow"),
                _dependency("ValidationRule", "Compte__c.Statut__c"),
            ]
        )

        assert usages == {
            "compte__c.statut__c": ["Flow", "Apex", "Validation Rule"]
        }

    def test_every_automation_kind_gets_a_readable_label(self) -> None:
        usages = field_automation_usages(
            [
                _dependency("trigger", "C__c.F__c"),
                _dependency("Formula", "C__c.F__c"),
                _dependency("Omni", "C__c.F__c"),
                _dependency("LWC", "C__c.F__c"),
                _dependency("Aura", "C__c.F__c"),
            ]
        )

        assert usages["c__c.f__c"] == [
            "Apex Trigger",
            "Formule",
            "OmniStudio",
            "LWC",
            "Aura",
        ]

    def test_presentation_kinds_are_ignored(self) -> None:
        """Nearly every field sits on a layout, so counting those would leave
        the column non-empty everywhere and warn about nothing."""
        usages = field_automation_usages(
            [
                _dependency("Layout", "Compte__c.Statut__c"),
                _dependency("FlexiPage", "Compte__c.Statut__c"),
                _dependency("Report", "Compte__c.Statut__c"),
            ]
        )

        assert usages == {}

    def test_object_and_apex_targets_are_ignored(self) -> None:
        usages = field_automation_usages(
            [
                Dependency(
                    source_name="MonFlow",
                    source_kind="Flow",
                    target_name="Compte__c",
                    target_kind="Object",
                ),
                Dependency(
                    source_name="MonFlow",
                    source_kind="Flow",
                    target_name="CompteService",
                    target_kind="Apex",
                ),
            ]
        )

        assert usages == {}


class TestAssignFieldAutomationUsages:
    def _snapshot(self, dependencies: list[Dependency]) -> MetadataSnapshot:
        return MetadataSnapshot(
            source_dir=Path("."),
            package_roots=[],
            objects=[
                ObjectInfo(
                    api_name="Compte__c",
                    fields=[
                        FieldInfo(api_name="Statut__c"),
                        FieldInfo(api_name="Inutilise__c"),
                    ],
                )
            ],
            dependencies=dependencies,
        )

    def test_fields_carry_their_own_usages(self) -> None:
        snapshot = self._snapshot([_dependency("Flow", "Compte__c.Statut__c")])

        assign_field_automation_usages(snapshot)

        statut, inutilise = snapshot.objects[0].fields
        assert statut.automation_usages == ["Flow"]
        assert statut.automation_usage_label == "Flow"
        assert inutilise.automation_usages == []
        assert inutilise.automation_usage_label == ""

    def test_matching_ignores_case(self) -> None:
        """Scans build the target name from various sources, so the field is
        matched case-insensitively rather than trusting an exact spelling."""
        snapshot = self._snapshot([_dependency("Flow", "compte__C.STATUT__c")])

        assign_field_automation_usages(snapshot)

        assert snapshot.objects[0].fields[0].automation_usages == ["Flow"]


OBJECT_META = """<?xml version="1.0" encoding="UTF-8"?>
<CustomObject xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Compte</label>
</CustomObject>
"""

FIELD_META = """<?xml version="1.0" encoding="UTF-8"?>
<CustomField xmlns="http://soap.sforce.com/2006/04/metadata">
    <fullName>{api_name}</fullName>
    <label>{api_name}</label>
    <type>Text</type>
</CustomField>
"""

FORMULA_FIELD_META = """<?xml version="1.0" encoding="UTF-8"?>
<CustomField xmlns="http://soap.sforce.com/2006/04/metadata">
    <fullName>Score__c</fullName>
    <label>Score</label>
    <type>Number</type>
    <formula>IF(ISBLANK(Remise__c), 0, Remise__c * 2)</formula>
</CustomField>
"""

VALIDATION_RULE_META = """<?xml version="1.0" encoding="UTF-8"?>
<ValidationRule xmlns="http://soap.sforce.com/2006/04/metadata">
    <fullName>Statut_Obligatoire</fullName>
    <active>true</active>
    <errorConditionFormula>ISBLANK(Statut__c)</errorConditionFormula>
    <errorMessage>Statut obligatoire</errorMessage>
</ValidationRule>
"""

FLOW_META = """<?xml version="1.0" encoding="UTF-8"?>
<Flow xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Maj Segment</label>
    <processType>AutoLaunchedFlow</processType>
    <status>Active</status>
    <start>
        <object>Compte__c</object>
        <triggerType>RecordAfterSave</triggerType>
    </start>
    <recordUpdates>
        <name>MajSegment</name>
        <label>Maj Segment</label>
        <inputAssignments>
            <field>Segment__c</field>
            <value><stringValue>Premium</stringValue></value>
        </inputAssignments>
    </recordUpdates>
</Flow>
"""

APEX_CLASS = """public class CompteService {
    public static void run() {
        List<Compte__c> comptes = [SELECT Id, Chiffre_Affaires__c FROM Compte__c];
        for (Compte__c c : comptes) {
            System.debug(c.Chiffre_Affaires__c);
        }
    }
}
"""


def _build_source(tmp_path: Path) -> Path:
    source = tmp_path / "source"
    object_dir = source / "objects" / "Compte__c"
    fields_dir = object_dir / "fields"
    fields_dir.mkdir(parents=True)
    (object_dir / "Compte__c.object-meta.xml").write_text(OBJECT_META, encoding="utf-8")
    for api_name in (
        "Statut__c",
        "Segment__c",
        "Chiffre_Affaires__c",
        "Remise__c",
        "Note_Interne__c",
    ):
        (fields_dir / f"{api_name}.field-meta.xml").write_text(
            FIELD_META.format(api_name=api_name), encoding="utf-8"
        )
    (fields_dir / "Score__c.field-meta.xml").write_text(
        FORMULA_FIELD_META, encoding="utf-8"
    )

    rules_dir = object_dir / "validationRules"
    rules_dir.mkdir()
    (rules_dir / "Statut_Obligatoire.validationRule-meta.xml").write_text(
        VALIDATION_RULE_META, encoding="utf-8"
    )

    flows_dir = source / "flows"
    flows_dir.mkdir()
    (flows_dir / "Maj_Segment.flow-meta.xml").write_text(FLOW_META, encoding="utf-8")

    classes_dir = source / "classes"
    classes_dir.mkdir()
    (classes_dir / "CompteService.cls").write_text(APEX_CLASS, encoding="utf-8")

    return source


class TestUsagesThroughTheFullParser:
    def _fields(self, tmp_path: Path) -> dict[str, FieldInfo]:
        snapshot = SalesforceMetadataParser(_build_source(tmp_path)).parse()
        return {
            field.api_name: field
            for obj in snapshot.objects
            for field in obj.fields
        }

    def test_each_source_kind_reaches_the_field_it_uses(self, tmp_path: Path) -> None:
        fields = self._fields(tmp_path)

        assert fields["Statut__c"].automation_usages == ["Validation Rule"]
        assert fields["Segment__c"].automation_usages == ["Flow"]
        assert fields["Chiffre_Affaires__c"].automation_usages == ["Apex"]
        assert fields["Remise__c"].automation_usages == ["Formule"]

    def test_an_unreferenced_field_stays_empty(self, tmp_path: Path) -> None:
        fields = self._fields(tmp_path)

        assert fields["Note_Interne__c"].automation_usage_label == ""

    def test_a_formula_field_is_not_a_usage_of_itself(self, tmp_path: Path) -> None:
        fields = self._fields(tmp_path)

        assert fields["Score__c"].automation_usage_label == ""
        assert fields["Score__c"].formula.startswith("IF(ISBLANK(Remise__c)")


COLUMN_HEADER = "Utilisé dans une automatisation ?"


class TestWorkbookColumn:
    def _fields_sheet(self, tmp_path: Path, **options: bool):
        account = ObjectInfo(
            api_name="Account",
            fields=[
                FieldInfo(
                    api_name="Statut__c",
                    description="Statut du compte",
                    automation_usages=["Flow", "Apex", "Validation Rule"],
                ),
                FieldInfo(api_name="Note__c", description="Note libre"),
            ],
        )
        paths = ExcelReportWriter().write_data_dictionary_workbooks(
            [account], tmp_path, **options
        )
        return load_workbook(paths[0])["Account"]

    def test_the_column_sits_between_description_and_the_dewey_comment(
        self, tmp_path: Path
    ) -> None:
        headers = [cell.value for cell in self._fields_sheet(tmp_path)[1]]

        index = headers.index(COLUMN_HEADER)
        assert headers[index - 1] == "Description"
        assert headers[index + 1] == "Commentaire Dewey"
        # Column H is the Description, so the new one is I.
        assert index + 1 == 9

    def test_usages_are_comma_separated_and_empty_when_unused(
        self, tmp_path: Path
    ) -> None:
        sheet = self._fields_sheet(tmp_path)
        headers = [cell.value for cell in sheet[1]]
        column = headers.index(COLUMN_HEADER) + 1

        assert sheet.cell(row=2, column=column).value == "Flow, Apex, Validation Rule"
        assert not sheet.cell(row=3, column=column).value

    def test_the_column_disappears_when_the_checkbox_is_cleared(
        self, tmp_path: Path
    ) -> None:
        sheet = self._fields_sheet(tmp_path, include_field_automation=False)
        headers = [cell.value for cell in sheet[1]]

        assert COLUMN_HEADER not in headers
        # The following columns close the gap instead of shifting.
        assert headers[headers.index("Description") + 1] == "Commentaire Dewey"
