"""Tests for the "Concatener la Description dans le Commentaire Dewey" toggle.

The toggle must apply to the object-level column (Synthese sheet) *and* to
the per-field column of each object sheet, each using its own Description.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.core.models import FieldInfo, ObjectInfo
from src.reporting.excel_writer import ExcelReportWriter


def _account() -> ObjectInfo:
    return ObjectInfo(
        api_name="Account",
        label="Compte",
        description="Description de l'objet",
        dewey_comment="Objet pivot",
        fields=[
            FieldInfo(
                api_name="Name",
                label="Nom",
                description="Description du champ",
                dewey_comment="Raison sociale",
            ),
            FieldInfo(api_name="Industry", label="Secteur", description="Secteur d'activite"),
        ],
    )


def _generate(tmp_path: Path, *, concat_description: bool) -> tuple[list, list]:
    writer = ExcelReportWriter()
    paths = writer.write_data_dictionary_workbooks(
        [_account()], tmp_path, concat_description=concat_description
    )
    workbook = load_workbook(paths[0])
    summary = workbook["Synthese"]
    fields_sheet = workbook["Account"]

    summary_headers = [cell.value for cell in summary[1]]
    object_comment = summary.cell(
        row=2, column=summary_headers.index("Commentaire Dewey") + 1
    ).value

    fields_headers = [cell.value for cell in fields_sheet[1]]
    comment_column = fields_headers.index("Commentaire Dewey") + 1
    field_comments = [
        fields_sheet.cell(row=row, column=comment_column).value for row in (2, 3)
    ]
    return object_comment, field_comments


def test_descriptions_are_concatenated_at_both_levels(tmp_path: Path) -> None:
    object_comment, field_comments = _generate(tmp_path, concat_description=True)

    assert object_comment == "Description de l'objet Objet pivot"
    assert field_comments == ["Description du champ Raison sociale", "Secteur d'activite"]


def test_raw_comments_are_kept_when_the_toggle_is_off(tmp_path: Path) -> None:
    object_comment, field_comments = _generate(tmp_path, concat_description=False)

    assert object_comment == "Objet pivot"
    assert field_comments == ["Raison sociale", None]
