"""Tests for the memory of the findings document.

Two behaviours are covered: a finding the analyzer stops reporting stays in
the cache and in the workbook with a "Terminé" status, and a workbook whose
columns were shifted still imports — rows included, findings and all, even
the ones Dewey does not know about.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from src.analyzer.engine import AnalyzerReport
from src.analyzer.models import Finding, Rule
from src.core.findings_cache import (
    adopt_findings,
    findings_cache_path,
    load_findings_cache,
    merge_history,
    save_findings_cache,
    write_findings_cache,
)
from src.core.findings_qualification import (
    RESOLVED_STATUS,
    FindingQualification,
    finding_keys,
)
from src.core.orchestrator.steps_mixin import _StepsMixin
from src.reporting.excel_reader_findings import (
    read_findings_qualifications,
    read_findings_workbook,
)
from src.reporting.excel_writer_findings import FINDINGS_SHEET, FindingsExcelWriter

RULE_ID = "APEX-SEC-003"


def _rule(rule_id: str = RULE_ID, severity: str = "Critical") -> Rule:
    return Rule(
        id=rule_id,
        enabled=True,
        scope="apex_class",
        category="Trusted",
        subcategory="Secure",
        severity=severity,
        source="PMD ApexSOQLInjection",
        reference="https://example.com/rule",
        title="Risque d'injection SOQL detecte",
        description="Requete dynamique sans echappement.",
        rationale="Un attaquant peut manipuler la requete.",
        remediation="Utiliser des variables de liaison.",
    )


def _finding(target_name: str, rule: Rule | None = None) -> Finding:
    return Finding(
        rule=rule or _rule(),
        target_kind="ApexClass",
        target_name=target_name,
        message="Risque d'injection SOQL dans une requete dynamique.",
        details=["Database.query() sans echappement."],
    )


def _qualification(**overrides) -> FindingQualification:
    defaults = dict(
        status="En cours",
        team="Squad CRM",
        target_sprint="S26-04",
        us_number="US-1234",
        us_title="Securiser AccountSelector",
        us_description="Remplacer la requete dynamique.",
        acceptance_criteria="Plus aucune Database.query() non echappee.",
    )
    defaults.update(overrides)
    return FindingQualification(**defaults)


def _report(*target_names: str) -> AnalyzerReport:
    rule = _rule()
    return AnalyzerReport(
        apex={name: [_finding(name, rule)] for name in target_names}
    )


class TestMergeHistory:
    def test_a_finding_that_disappeared_is_kept_and_flagged(self) -> None:
        findings, resolved = merge_history(
            [_finding("Alpha")], [_finding("Alpha"), _finding("Beta")]
        )

        assert [finding.target_name for finding in findings] == ["Alpha", "Beta"]
        assert resolved == {("Beta", RULE_ID, 0)}

    def test_a_finding_detected_again_is_no_longer_flagged(self) -> None:
        findings, resolved = merge_history(
            [_finding("Alpha"), _finding("Beta")], [_finding("Beta")]
        )

        assert len(findings) == 2
        assert resolved == set()

    def test_nothing_known_yet_flags_nothing(self) -> None:
        findings, resolved = merge_history([_finding("Alpha")], [])

        assert [finding.target_name for finding in findings] == ["Alpha"]
        assert resolved == set()

    def test_the_last_occurrence_of_a_pair_is_the_one_flagged(self) -> None:
        """Losing one of three identical findings must not renumber the two
        that remain, or their qualifications would move to another row."""
        findings, resolved = merge_history(
            [_finding("Alpha"), _finding("Alpha")],
            [_finding("Alpha"), _finding("Alpha"), _finding("Alpha")],
        )

        assert len(findings) == 3
        assert resolved == {("Alpha", RULE_ID, 2)}

    def test_an_org_with_no_finding_left_keeps_its_whole_past(self) -> None:
        findings, resolved = merge_history([], [_finding("Alpha"), _finding("Beta")])

        assert len(findings) == 2
        assert resolved == {("Alpha", RULE_ID, 0), ("Beta", RULE_ID, 0)}


class TestAdoptFindings:
    def test_an_added_occurrence_lines_up_with_the_key_of_its_file_row(self) -> None:
        """The invariant the import relies on: a finding read from a file and
        added to the known ones takes the very key its row had."""
        rule = _rule()
        known = [_finding("Alpha", rule), _finding("Alpha", rule)]
        added = _finding("Alpha", rule)

        merged = adopt_findings(known, [added])

        assert finding_keys(merged) == [("Alpha", RULE_ID, index) for index in range(3)]
        assert merged[2] is added

    def test_a_rule_dewey_still_knows_is_preferred_to_the_rebuilt_one(self) -> None:
        """A workbook exports neither the scope nor the API range of a rule,
        so the addition would otherwise carry a degraded copy of it."""
        known = [_finding("Alpha")]
        rebuilt = Rule(
            id=RULE_ID,
            enabled=True,
            scope="",
            category="Trusted",
            subcategory="Secure",
            severity="Info",
            source="",
            reference="",
            title="",
            description="",
            rationale="",
            remediation="",
        )

        merged = adopt_findings(known, [_finding("Beta", rebuilt)])

        assert merged[1].rule is known[0].rule
        assert merged[1].rule.scope == "apex_class"

    def test_an_unknown_rule_keeps_what_the_file_said(self) -> None:
        addition = _finding("Gamma", _rule("FLOW-DOC-001", "Minor"))

        merged = adopt_findings([_finding("Alpha")], [addition])

        assert merged[1].rule.id == "FLOW-DOC-001"
        assert merged[1].rule.severity == "Minor"

    def test_adding_nothing_leaves_the_known_findings_in_place(self) -> None:
        known = [_finding("Beta"), _finding("Alpha")]

        assert [
            finding.target_name for finding in adopt_findings(known, [])
        ] == ["Alpha", "Beta"]


class TestCacheRemembersThePast:
    def test_a_second_run_keeps_the_findings_of_the_first(self, tmp_path: Path) -> None:
        path = findings_cache_path(tmp_path, "MHINT")
        save_findings_cache(_report("Alpha", "Beta"), path, alias="MHINT")
        save_findings_cache(_report("Alpha"), path, alias="MHINT")

        cached = load_findings_cache(path)

        assert cached is not None
        assert [finding.target_name for finding in cached.findings] == ["Alpha", "Beta"]
        assert cached.resolved_keys == {("Beta", RULE_ID, 0)}

    def test_a_finding_coming_back_clears_its_flag(self, tmp_path: Path) -> None:
        path = findings_cache_path(tmp_path, "MHINT")
        save_findings_cache(_report("Alpha", "Beta"), path, alias="MHINT")
        save_findings_cache(_report("Alpha"), path, alias="MHINT")
        save_findings_cache(_report("Alpha", "Beta"), path, alias="MHINT")

        cached = load_findings_cache(path)

        assert cached is not None
        assert len(cached.findings) == 2
        assert cached.resolved_keys == set()

    def test_a_cache_written_before_the_history_existed_flags_nothing(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "legacy.json"
        path.write_text(
            '{"version": 1, "alias": "MHINT", "findings": ['
            '{"rule": {"id": "R1", "enabled": true, "scope": "", "category": "",'
            ' "subcategory": "", "severity": "Major", "source": "", "reference": "",'
            ' "title": "", "description": "", "rationale": "", "remediation": ""},'
            ' "target_kind": "ApexClass", "target_name": "Alpha", "message": "",'
            ' "details": []}]}',
            encoding="utf-8",
        )

        cached = load_findings_cache(path)

        assert cached is not None
        assert cached.resolved_keys == set()

    def test_added_findings_can_be_stored_without_a_run(self, tmp_path: Path) -> None:
        path = findings_cache_path(tmp_path, "MHINT")
        save_findings_cache(_report("Alpha"), path, alias="MHINT")

        write_findings_cache(
            [_finding("Alpha"), _finding("Zeta")],
            path,
            alias="MHINT",
            resolved_keys={("Zeta", RULE_ID, 0)},
        )
        cached = load_findings_cache(path)

        assert cached is not None
        assert [finding.target_name for finding in cached.findings] == ["Alpha", "Zeta"]
        assert cached.resolved_keys == {("Zeta", RULE_ID, 0)}


class _GenerationStep(_StepsMixin):
    """Just enough of the generator to exercise its findings step."""

    def __init__(self, history_path: Path | None) -> None:
        self.findings_history_path = history_path
        self.log = lambda message: None


class TestGenerationReadsTheHistory:
    def test_the_document_keeps_the_findings_of_the_previous_run(
        self, tmp_path: Path
    ) -> None:
        path = findings_cache_path(tmp_path, "MHINT")
        save_findings_cache(_report("Alpha", "Beta"), path, alias="MHINT")

        findings, resolved = _GenerationStep(path)._findings_with_history(
            _report("Alpha")
        )

        assert [finding.target_name for finding in findings] == ["Alpha", "Beta"]
        assert resolved == {("Beta", RULE_ID, 0)}

    def test_without_a_history_the_run_stands_on_its_own(self) -> None:
        findings, resolved = _GenerationStep(None)._findings_with_history(
            _report("Alpha")
        )

        assert [finding.target_name for finding in findings] == ["Alpha"]
        assert resolved == set()


class TestResolvedRowsInTheWorkbook:
    def test_a_resolved_finding_is_exported_with_the_resolved_status(
        self, tmp_path: Path
    ) -> None:
        key = ("Beta", RULE_ID, 0)
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding("Alpha"), _finding("Beta")],
            tmp_path / "f.xlsx",
            qualifications={key: _qualification()},
            resolved_keys={key},
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert sheet["E4"].value == "Beta"
        assert sheet["M4"].value == RESOLVED_STATUS
        # The rest of the TechLead's work is left alone.
        assert sheet["N4"].value == "Squad CRM"
        assert sheet["P4"].value == "US-1234"
        assert sheet["M3"].value is None

    def test_a_resolved_finding_without_a_qualification_only_gets_the_status(
        self, tmp_path: Path
    ) -> None:
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding("Alpha")],
            tmp_path / "f.xlsx",
            resolved_keys={("Alpha", RULE_ID, 0)},
        )

        sheet = openpyxl.load_workbook(path)[FINDINGS_SHEET]
        assert sheet["M3"].value == RESOLVED_STATUS
        assert [sheet.cell(row=3, column=col).value for col in range(14, 20)] == [
            None
        ] * 6

    def test_the_resolved_status_is_read_back_on_import(self, tmp_path: Path) -> None:
        key = ("Alpha", RULE_ID, 0)
        path = FindingsExcelWriter().write_findings_workbook(
            [_finding("Alpha")], tmp_path / "f.xlsx", resolved_keys={key}
        )

        assert read_findings_qualifications(path)[key].status == RESOLVED_STATUS


class TestImportFollowsTheColumns:
    def _export(self, tmp_path: Path) -> Path:
        return FindingsExcelWriter().write_findings_workbook(
            [_finding("Alpha"), _finding("Beta")],
            tmp_path / "f.xlsx",
            qualifications={
                ("Alpha", RULE_ID, 0): _qualification(us_number="US-1"),
                ("Beta", RULE_ID, 0): _qualification(us_number="US-2"),
            },
        )

    def test_an_inserted_column_does_not_shift_the_values(
        self, tmp_path: Path
    ) -> None:
        path = self._export(tmp_path)
        workbook = openpyxl.load_workbook(path)
        workbook[FINDINGS_SHEET].insert_cols(1)
        workbook.save(path)

        imported = read_findings_qualifications(path)

        assert imported[("Alpha", RULE_ID, 0)].us_number == "US-1"
        assert imported[("Beta", RULE_ID, 0)].us_number == "US-2"

    def test_a_removed_column_does_not_shift_the_values(self, tmp_path: Path) -> None:
        path = self._export(tmp_path)
        workbook = openpyxl.load_workbook(path)
        workbook[FINDINGS_SHEET].delete_cols(2)  # "Catégorie"
        workbook.save(path)

        imported = read_findings_qualifications(path)

        assert imported[("Alpha", RULE_ID, 0)].team == "Squad CRM"
        assert imported[("Beta", RULE_ID, 0)].us_number == "US-2"

    def test_every_row_yields_the_finding_it_describes(self, tmp_path: Path) -> None:
        rows = read_findings_workbook(self._export(tmp_path))

        assert [row.key for row in rows] == [
            ("Alpha", RULE_ID, 0),
            ("Beta", RULE_ID, 0),
        ]
        finding = rows[0].finding
        assert finding.rule.id == RULE_ID
        assert finding.rule.severity == "Critical"
        assert finding.rule.category == "Trusted"
        assert finding.rule.subcategory == "Secure"
        assert finding.rule.title == "Risque d'injection SOQL detecte"
        assert finding.rule.remediation == "Utiliser des variables de liaison."
        assert finding.target_kind == "ApexClass"
        assert finding.target_name == "Alpha"
        assert finding.details == ["Database.query() sans echappement."]

    def test_a_row_dewey_does_not_know_about_is_readable_too(
        self, tmp_path: Path
    ) -> None:
        """What the screen relies on to take an unknown row in instead of
        dropping it: the row carries a usable finding of its own."""
        path = self._export(tmp_path)
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[FINDINGS_SHEET]
        sheet["A5"] = "Majeur"
        sheet["C5"] = "FLOW-DOC-001"
        sheet["D5"] = "Flow"
        sheet["E5"] = "Gamma_Flow"
        sheet["M5"] = "Faux positif"
        workbook.save(path)

        rows = read_findings_workbook(path)

        assert len(rows) == 3
        unknown = rows[2]
        assert unknown.key == ("Gamma_Flow", "FLOW-DOC-001", 0)
        assert unknown.qualification.status == "Faux positif"
        assert unknown.finding.rule.severity == "Major"
        assert unknown.finding.target_kind == "Flow"
        assert unknown.finding.target_name == "Gamma_Flow"
