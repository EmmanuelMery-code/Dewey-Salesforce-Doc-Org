"""Tests: "Documentation > Creer les CSV des picklists" export.

Contracts tested:
  - FieldInfo.picklist_values / picklist_api_names are parallel lists
    (label, API name) for both local and Global Value Set picklists.
  - PicklistCsvWriter.write_picklist_csv_export() creates
    picklist/{fields,global}/*.csv (2 columns: Label, Nom API) plus a
    picklist/picklists_summary.xlsx summary workbook with the columns
    Object | Champs | Nom de la Global Picklist | Repertoire fichier | Nom Fichier.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from src.parsers.salesforce_parser import SalesforceMetadataParser
from src.reporting.picklist_csv_writer import PicklistCsvWriter

ACCOUNT_OBJECT_META = """<?xml version="1.0" encoding="UTF-8"?>
<CustomObject xmlns="http://soap.sforce.com/2006/04/metadata">
    <label>Account</label>
</CustomObject>
"""

LOCAL_PICKLIST_FIELD = """<?xml version="1.0" encoding="UTF-8"?>
<CustomField xmlns="http://soap.sforce.com/2006/04/metadata">
    <fullName>Status__c</fullName>
    <label>Status</label>
    <type>Picklist</type>
    <valueSet>
        <valueSetDefinition>
            <value>
                <fullName>New</fullName>
                <label>Nouveau</label>
            </value>
            <value>
                <fullName>Closed</fullName>
                <label>Ferme</label>
            </value>
        </valueSetDefinition>
    </valueSet>
</CustomField>
"""

GLOBAL_PICKLIST_FIELD = """<?xml version="1.0" encoding="UTF-8"?>
<CustomField xmlns="http://soap.sforce.com/2006/04/metadata">
    <fullName>Priority__c</fullName>
    <label>Priority</label>
    <type>Picklist</type>
    <valueSet>
        <valueSetName>PriorityPicklist</valueSetName>
    </valueSet>
</CustomField>
"""

GLOBAL_VALUE_SET = """<?xml version="1.0" encoding="UTF-8"?>
<GlobalValueSet xmlns="http://soap.sforce.com/2006/04/metadata">
    <customValue>
        <fullName>HIGH</fullName>
        <label>Haute</label>
    </customValue>
    <customValue>
        <fullName>LOW</fullName>
        <label>Basse</label>
    </customValue>
</GlobalValueSet>
"""


def _build_source(tmp_path: Path) -> Path:
    source = tmp_path / "source"
    fields_dir = source / "objects" / "Account" / "fields"
    fields_dir.mkdir(parents=True, exist_ok=True)
    (source / "objects" / "Account" / "Account.object-meta.xml").write_text(
        ACCOUNT_OBJECT_META, encoding="utf-8"
    )
    (fields_dir / "Status__c.field-meta.xml").write_text(LOCAL_PICKLIST_FIELD, encoding="utf-8")
    (fields_dir / "Priority__c.field-meta.xml").write_text(GLOBAL_PICKLIST_FIELD, encoding="utf-8")

    gvs_dir = source / "globalValueSets"
    gvs_dir.mkdir(parents=True, exist_ok=True)
    (gvs_dir / "PriorityPicklist.globalValueSet-meta.xml").write_text(
        GLOBAL_VALUE_SET, encoding="utf-8"
    )
    return source


def _read_csv_rows(path: Path) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        return list(csv.reader(handle, delimiter=";"))


class TestFieldInfoPicklistLabelsAndApiNames:
    def test_local_picklist_has_parallel_label_and_api_name_lists(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        snapshot = SalesforceMetadataParser(source).parse()
        account = next(obj for obj in snapshot.objects if obj.api_name == "Account")
        status = next(f for f in account.fields if f.api_name == "Status__c")

        assert status.is_picklist
        assert status.picklist_is_global is False
        assert status.picklist_values == ["Nouveau", "Ferme"]
        assert status.picklist_api_names == ["New", "Closed"]

    def test_global_picklist_resolves_labels_and_api_names_from_gvs(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        snapshot = SalesforceMetadataParser(source).parse()
        account = next(obj for obj in snapshot.objects if obj.api_name == "Account")
        priority = next(f for f in account.fields if f.api_name == "Priority__c")

        assert priority.picklist_is_global is True
        assert priority.picklist_global_name == "PriorityPicklist"
        assert priority.picklist_values == ["Haute", "Basse"]
        assert priority.picklist_api_names == ["HIGH", "LOW"]


class TestPicklistCsvWriter:
    def test_creates_expected_folder_tree(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        snapshot = SalesforceMetadataParser(source).parse()
        output_dir = tmp_path / "output"

        writer = PicklistCsvWriter()
        summary_path = writer.write_picklist_csv_export(snapshot.objects, output_dir)

        picklist_root = output_dir / "picklist"
        assert (picklist_root / "fields").is_dir()
        assert (picklist_root / "global").is_dir()
        assert summary_path == picklist_root / "picklists_summary.xlsx"
        assert summary_path.exists()

    def test_local_picklist_csv_has_label_and_api_name_columns(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        snapshot = SalesforceMetadataParser(source).parse()
        output_dir = tmp_path / "output"

        PicklistCsvWriter().write_picklist_csv_export(snapshot.objects, output_dir)

        csv_path = output_dir / "picklist" / "fields" / "Account_Status__c.csv"
        assert csv_path.exists()
        rows = _read_csv_rows(csv_path)
        assert rows[0] == ["Label", "Nom API"]
        assert rows[1:] == [["Nouveau", "New"], ["Ferme", "Closed"]]

    def test_global_picklist_csv_is_named_after_the_global_value_set(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        snapshot = SalesforceMetadataParser(source).parse()
        output_dir = tmp_path / "output"

        PicklistCsvWriter().write_picklist_csv_export(snapshot.objects, output_dir)

        csv_path = output_dir / "picklist" / "global" / "PriorityPicklist.csv"
        assert csv_path.exists()
        rows = _read_csv_rows(csv_path)
        assert rows[0] == ["Label", "Nom API"]
        assert rows[1:] == [["Haute", "HIGH"], ["Basse", "LOW"]]

        # A local picklist file must not be created for a global picklist.
        assert not (output_dir / "picklist" / "fields" / "Account_Priority__c.csv").exists()

    def test_summary_workbook_lists_every_picklist_field(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        snapshot = SalesforceMetadataParser(source).parse()
        output_dir = tmp_path / "output"

        summary_path = PicklistCsvWriter().write_picklist_csv_export(snapshot.objects, output_dir)

        workbook = load_workbook(summary_path)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        assert rows[0] == (
            "Object",
            "Champs",
            "Nom de la Global Picklist",
            "Repertoire fichier",
            "Nom Fichier",
        )
        data_rows = {row[:2]: row for row in rows[1:]}
        assert data_rows[("Account", "Status__c")] == (
            "Account",
            "Status__c",
            "-",
            "fields",
            "Account_Status__c.csv",
        )
        assert data_rows[("Account", "Priority__c")] == (
            "Account",
            "Priority__c",
            "PriorityPicklist",
            "global",
            "PriorityPicklist.csv",
        )

    def test_shared_global_value_set_is_written_only_once(self, tmp_path: Path) -> None:
        source = _build_source(tmp_path)
        # Add a second field on the same object referencing the same GVS.
        second_field = GLOBAL_PICKLIST_FIELD.replace("Priority__c", "Priority2__c")
        (source / "objects" / "Account" / "fields" / "Priority2__c.field-meta.xml").write_text(
            second_field, encoding="utf-8"
        )
        snapshot = SalesforceMetadataParser(source).parse()
        output_dir = tmp_path / "output"

        PicklistCsvWriter().write_picklist_csv_export(snapshot.objects, output_dir)

        global_files = list((output_dir / "picklist" / "global").glob("*.csv"))
        assert len(global_files) == 1
