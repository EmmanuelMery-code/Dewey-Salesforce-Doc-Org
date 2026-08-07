import os
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_global_value_sets(root_directory, ns):
    """
    Recherche les fichiers Global Value Sets (.globalValueSet-meta.xml)
    et charge le dictionnaire : { nom_globalValueSet : [valeur1, valeur2, ...] }
    """
    global_value_sets = {}
    
    for root, dirs, files in os.walk(root_directory):
        if os.path.basename(root).lower() == 'globalvaluesets':
            for file_name in files:
                if file_name.endswith('.xml'):
                    gvs_name = file_name.split('.')[0]
                    file_path = os.path.join(root, file_name)
                    
                    try:
                        tree = ET.parse(file_path)
                        xml_root = tree.getroot()
                        
                        def find_elements_all(parent, element_name):
                            res = parent.findall(f'sf:{element_name}', ns)
                            if not res:
                                res = parent.findall(element_name)
                            return res

                        values_list = []
                        val_nodes = find_elements_all(xml_root, 'customValue')
                        for val in val_nodes:
                            lbl = val.find('sf:label', ns) if val.find('sf:label', ns) is not None else val.find('label')
                            fn = val.find('sf:fullName', ns) if val.find('sf:fullName', ns) is not None else val.find('fullName')
                            
                            val_text = ""
                            if lbl is not None and lbl.text:
                                val_text = lbl.text
                            elif fn is not None and fn.text:
                                val_text = fn.text

                            if val_text:
                                values_list.append(val_text)

                        global_value_sets[gvs_name] = values_list
                    except Exception as e:
                        print(f"Erreur lors de la lecture du GlobalValueSet {file_path}: {e}")
                        
    return global_value_sets


def extract_picklist_fields(root_directory):
    """
    Parcourt les sous-répertoires, recherche les fichiers XML dans les dossiers 'fields'
    et extrait les métadonnées des champs de type Picklist / MultiselectPicklist.
    """
    data_rows = []
    ns = {'sf': 'http://soap.sforce.com/2006/04/metadata'}

    # 1. Charger d'abord les Global Value Sets s'ils sont présents dans le projet
    global_value_sets = load_global_value_sets(root_directory, ns)

    for root, dirs, files in os.walk(root_directory):
        if os.path.basename(root).lower() == 'fields':
            object_name = os.path.basename(os.path.dirname(root))
            
            for file_name in files:
                if file_name.endswith('.xml'):
                    file_path = os.path.join(root, file_name)
                    
                    try:
                        tree = ET.parse(file_path)
                        xml_root = tree.getroot()
                        
                        def find_element(parent, element_name):
                            el = parent.find(f'sf:{element_name}', ns)
                            if el is None:
                                el = parent.find(element_name)
                            return el

                        def find_elements_all(parent, element_name):
                            res = parent.findall(f'sf:{element_name}', ns)
                            if not res:
                                res = parent.findall(element_name)
                            return res

                        # Vérification du type de champ
                        type_node = find_element(xml_root, 'type')
                        if type_node is not None and type_node.text in ['Picklist', 'MultiselectPicklist']:
                            field_type = type_node.text
                            
                            fullname_node = find_element(xml_root, 'fullName')
                            if fullname_node is not None and fullname_node.text:
                                field_name = fullname_node.text
                            else:
                                field_name = file_name.replace('.field-meta.xml', '').replace('.xml', '')

                            is_global = "Non"
                            global_name = "-"
                            values_list = []

                            value_set_node = find_element(xml_root, 'valueSet')
                            if value_set_node is not None:
                                global_name_node = find_element(value_set_node, 'valueSetName')

                                # Cas 1: Picklist globale
                                if global_name_node is not None and global_name_node.text:
                                    is_global = "Oui"
                                    global_name = global_name_node.text
                                    
                                    # Si le fichier .globalValueSet correspondant a été trouvé
                                    if global_name in global_value_sets:
                                        values_list = global_value_sets[global_name]

                                # Cas 2: Valeurs locales du champ
                                val_def_node = find_element(value_set_node, 'valueSetDefinition')
                                target_parent = val_def_node if val_def_node is not None else value_set_node
                                
                                val_nodes = find_elements_all(target_parent, 'value')
                                for val in val_nodes:
                                    lbl = find_element(val, 'label')
                                    fn = find_element(val, 'fullName')
                                    
                                    val_text = ""
                                    if lbl is not None and lbl.text:
                                        val_text = lbl.text
                                    elif fn is not None and fn.text:
                                        val_text = fn.text

                                    if val_text and val_text not in values_list:
                                        values_list.append(val_text)

                            # Formate les valeurs séparées par ' | '
                            values_str = " | ".join(values_list) if values_list else "-"

                            data_rows.append({
                                'object': object_name,
                                'field': field_name,
                                'type': field_type,
                                'is_global': is_global,
                                'global_name': global_name,
                                'values': values_str
                            })

                    except Exception as e:
                        print(f"Erreur d'analyse sur le fichier {file_path}: {e}")

    return data_rows


def generate_excel_report(data_rows, output_filename="champsPickList.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Champs Picklist"

    ws.views.sheetView[0].showGridLines = True

    COLOR_NAVY = "1F4E78"
    COLOR_ZEBRA = "F2F5F9"
    COLOR_BORDER = "D9D9D9"

    font_title = Font(name="Calibri", size=16, bold=True, color=COLOR_NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)

    fill_header = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )

    ws["A1"] = "Rapport des Champs Picklist (Salesforce Metadata)"
    ws["A1"].font = font_title
    ws["A2"] = f"Total des champs trouvés : {len(data_rows)}"
    ws["A2"].font = font_subtitle

    headers = [
        "Nom de l'Objet",
        "Nom du Champ",
        "Type de Champ",
        "Picklist Globale ?",
        "Nom Picklist Globale",
        "Valeurs de la Picklist"
    ]

    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    start_row = 5
    for row_idx, data in enumerate(data_rows, start=start_row):
        row_fill = fill_zebra if row_idx % 2 == 0 else None

        row_data = [
            data['object'],
            data['field'],
            data['type'],
            data['is_global'],
            data['global_name'],
            data['values']
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            
            if row_fill:
                cell.fill = row_fill

            if col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_idx == 4 and val == "Oui":
                cell.font = font_bold

    ws.freeze_panes = "A5"

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row < header_row:
                continue
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    wb.save(output_filename)
    print(f"Fichier Excel créé : {output_filename}")


if __name__ == "__main__":
    DOSSIER_RACINE = "C:/Users/emery/Desktop/Dewey/org/test new version/retrieve/force-app/main/default/objects" # Remplacez par votre dossier si nécessaire
    donnees = extract_picklist_fields(DOSSIER_RACINE)
    generate_excel_report(donnees, "champsPickList.xlsx")