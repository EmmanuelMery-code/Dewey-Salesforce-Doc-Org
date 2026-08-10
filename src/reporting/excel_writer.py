from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.models import ObjectInfo, PmdViolation, SecurityArtifact
from src.reporting.excel_writer_data_dictionary import _ExcelDataDictionaryMixin

LogCallback = Callable[[str], None]


class ExcelReportWriter(_ExcelDataDictionaryMixin):
    """Produce ``.xlsx`` documents from the parsed metadata snapshot.

    Each ``write_*`` method writes one workbook (security, inventory, data
    dictionary, PMD violations, ...) and returns the resulting path so the
    orchestrator can collect all artefacts in :class:`GenerationResult`.
    Data Dictionary generation itself lives in
    :class:`~src.reporting.excel_writer_data_dictionary._ExcelDataDictionaryMixin`.
    """

    def __init__(self, log_callback: LogCallback | None = None) -> None:
        self.log: LogCallback = log_callback or (lambda message: None)

    def write_security_workbook(
        self, artifacts: list[SecurityArtifact], output_path: str | Path, title: str
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Synthese"
        self._write_sheet(
            summary,
            [
                "Nom",
                "Label",
                "Description",
                "Nb droits objet",
                "Nb droits champ",
                "Nb permissions systeme",
                "Nb applis",
                "Nb onglets",
                "Nb classes",
                "Nb flows",
            ],
            [
                [
                    artifact.name,
                    artifact.label,
                    artifact.description,
                    len(artifact.object_permissions),
                    len(artifact.field_permissions),
                    len(artifact.user_permissions),
                    len(artifact.application_visibilities),
                    len(artifact.tab_visibilities),
                    len(artifact.class_accesses),
                    len(artifact.flow_accesses),
                ]
                for artifact in artifacts
            ],
        )

        self._write_sheet(
            workbook.create_sheet("DroitsObjet"),
            ["Nom", "Objet", "Lecture", "Creation", "Modification", "Suppression", "ViewAll", "ModifyAll"],
            [
                [
                    artifact.name,
                    permission.object_name,
                    permission.allow_read,
                    permission.allow_create,
                    permission.allow_edit,
                    permission.allow_delete,
                    permission.view_all_records,
                    permission.modify_all_records,
                ]
                for artifact in artifacts
                for permission in artifact.object_permissions
            ],
        )

        self._write_sheet(
            workbook.create_sheet("DroitsChamp"),
            ["Nom", "Champ", "Lecture", "Modification"],
            [
                [artifact.name, permission.field_name, permission.readable, permission.editable]
                for artifact in artifacts
                for permission in artifact.field_permissions
            ],
        )

        self._write_sheet(
            workbook.create_sheet("PermissionsSysteme"),
            ["Nom", "Permission", "Activee"],
            [
                [artifact.name, permission.name, permission.enabled]
                for artifact in artifacts
                for permission in artifact.user_permissions
            ],
        )

        self._write_sheet(
            workbook.create_sheet("Applications"),
            ["Nom", "Application", "Visible", "Defaut"],
            [
                [artifact.name, app.name, app.visible, app.default]
                for artifact in artifacts
                for app in artifact.application_visibilities
            ],
        )

        self._write_sheet(
            workbook.create_sheet("Onglets"),
            ["Nom", "Onglet", "Visibilite", "Defaut"],
            [
                [artifact.name, tab.name, tab.visible, tab.default]
                for artifact in artifacts
                for tab in artifact.tab_visibilities
            ],
        )

        self._write_sheet(
            workbook.create_sheet("ClassesApex"),
            ["Nom", "Classe Apex", "Active"],
            [
                [artifact.name, access.name, access.enabled]
                for artifact in artifacts
                for access in artifact.class_accesses
            ],
        )

        self._write_sheet(
            workbook.create_sheet("Flows"),
            ["Nom", "Flow", "Actif"],
            [
                [artifact.name, access.name, access.enabled]
                for artifact in artifacts
                for access in artifact.flow_accesses
            ],
        )

        self._write_sheet(
            workbook.create_sheet("RecordTypes"),
            ["Nom", "Record Type", "Visible", "Defaut"],
            [
                [artifact.name, item.record_type, item.visible, item.default]
                for artifact in artifacts
                for item in artifact.record_type_visibilities
            ],
        )

        self._write_sheet(
            workbook.create_sheet("CustomPermissions"),
            ["Nom", "Custom Permission", "Activee"],
            [
                [artifact.name, access.name, access.enabled]
                for artifact in artifacts
                for access in artifact.custom_permissions
            ],
        )

        workbook.save(output)
        self.log(f"{title} genere: {output}")
        return output

    def write_inventory_workbook(
        self, inventory: dict[str, list[dict[str, object]]], output_path: str | Path
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet_definitions = [
            ("RecordTypes", "record_types"),
            ("Layouts", "layouts"),
            ("LightningPages", "lightning_pages"),
            ("ValidationRules", "validation_rules"),
            ("OmniStudio", "omnistudio"),
            ("BusinessRules", "business_rules_engine"),
            ("Flows", "flows"),
            ("PermissionSets", "permission_sets"),
            ("Profiles", "profiles"),
            ("Reports", "reports"),
            ("Dashboards", "dashboards"),
        ]

        first_title, first_key = sheet_definitions[0]
        first_rows = inventory.get(first_key, [])
        summary = workbook.active
        summary.title = first_title
        self._write_dict_sheet(summary, first_rows)

        for title, key in sheet_definitions[1:]:
            self._write_dict_sheet(workbook.create_sheet(title), inventory.get(key, []))

        workbook.save(output)
        self.log(f"Classeur inventaire metadata genere: {output}")
        return output

    def write_pmd_workbook(
        self,
        violations_by_artifact: dict[str, list[PmdViolation]],
        output_path: str | Path,
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Synthese"
        self._write_sheet(
            summary,
            ["Composant", "Violations PMD"],
            [
                [artifact_name, len(violations)]
                for artifact_name, violations in sorted(
                    violations_by_artifact.items(), key=lambda item: item[0].lower()
                )
            ],
        )

        detail_rows = [
            [
                artifact_name,
                violation.rule,
                violation.ruleset,
                violation.priority,
                violation.begin_line,
                violation.end_line,
                violation.message,
                str(violation.file_path),
            ]
            for artifact_name, violations in sorted(
                violations_by_artifact.items(), key=lambda item: item[0].lower()
            )
            for violation in violations
        ]
        self._write_sheet(
            workbook.create_sheet("Violations"),
            ["Composant", "Regle", "Ruleset", "Priorite", "LigneDebut", "LigneFin", "Message", "Fichier"],
            detail_rows,
        )

        workbook.save(output)
        self.log(f"Classeur PMD genere: {output}")
        return output

    def write_picklists_workbook(
        self, objects: list[ObjectInfo], output_path: str | Path
    ) -> Path:
        """Generate the Picklist fields inventory workbook.

        Lists every ``Picklist``/``MultiselectPicklist`` field found across
        ``objects``, resolving Global Value Set references to their values
        (already computed by the parser onto ``FieldInfo.picklist_values``).
        """
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        rows = [
            [
                obj.api_name,
                item.api_name,
                item.data_type,
                "Oui" if item.picklist_is_global else "Non",
                item.picklist_global_name or "-",
                " | ".join(item.picklist_values) if item.picklist_values else "-",
            ]
            for obj in objects
            for item in obj.fields
            if item.data_type in ("Picklist", "MultiselectPicklist")
        ]

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Champs Picklist"
        self._write_sheet(
            summary,
            [
                "Nom de l'Objet",
                "Nom du Champ",
                "Type de Champ",
                "Picklist Globale ?",
                "Nom Picklist Globale",
                "Valeurs de la Picklist",
            ],
            rows,
        )

        workbook.save(output)
        self.log(f"Classeur Picklist genere ({len(rows)} champ(s)): {output}")
        return output

    def _write_sheet(self, worksheet, headers: list[str], rows: list[list[object]]) -> None:
        worksheet.append(headers)
        header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in rows:
            worksheet.append(row)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for index, header in enumerate(headers, start=1):
            max_length = len(header)
            for row in worksheet.iter_rows(min_col=index, max_col=index, min_row=2):
                value = row[0].value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 60)

    def _write_dict_sheet(self, worksheet, rows: list[dict[str, object]]) -> None:
        if rows:
            headers = list(rows[0].keys())
            data = [[row.get(header, "") for header in headers] for row in rows]
        else:
            headers = ["Information"]
            data = [["Aucune donnee trouvee"]]
        self._write_sheet(worksheet, headers, data)
