"""Read a reviewed findings workbook back into Dewey.

Counterpart of :mod:`src.reporting.excel_writer_findings`. Each data row
yields both the TechLead columns (Qualification and US) and the finding they
describe, so a row Dewey does not know about — a finding dropped by the
analyzer, or one the TechLead added by hand — can be taken in instead of
being discarded.

Columns are located from the header row rather than from fixed letters: a
file whose columns were shifted, because one was inserted or removed, still
imports as long as the headers Dewey wrote are recognisable. The fixed
``A..S`` layout is only the fallback for a file with no usable header row.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.analyzer.models import Finding, Rule
from src.core.findings_qualification import (
    QUALIFICATION_FIELD_COUNT,
    FindingQualification,
    QualificationKey,
    assign_keys,
)
from src.reporting.excel_writer_findings import (
    COLUMN_FIELDS,
    FINDINGS_SHEET,
    FIRST_DATA_ROW,
    HEADERS,
    SEVERITY_LABELS,
)

#: How far down the sheet the header row is looked for, and how many headers
#: a row must name to be taken for it. Four is low enough for a workbook
#: whose columns were partly renamed, high enough not to match a data row.
_HEADER_SEARCH_ROWS = 10
_MIN_RECOGNISED_HEADERS = 4

_QUALIFICATION_FIELDS = COLUMN_FIELDS[-QUALIFICATION_FIELD_COUNT:]
_SEVERITY_CODES = {label: code for code, label in SEVERITY_LABELS.items()}

_NON_ALPHANUMERIC_RE = re.compile(r"[^a-z0-9]+")


class FindingsWorkbookError(Exception):
    """Raised when a file cannot be read as a Dewey findings workbook."""


@dataclass(slots=True)
class ImportedFinding:
    """One data row of a reviewed workbook."""

    key: QualificationKey
    qualification: FindingQualification
    finding: Finding


def read_findings_workbook(workbook_path: str | Path) -> list[ImportedFinding]:
    """Every data row of ``workbook_path``, in file order.

    Rows whose component and rule are both blank are skipped (trailing
    formatting leftovers). Occurrence indexes follow the row order of the
    file, so an untouched Dewey export re-imports onto the exact same
    findings.
    """
    path = Path(workbook_path)
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except OSError as exc:
        raise FindingsWorkbookError(str(exc)) from exc
    except Exception as exc:  # openpyxl raises assorted types on bad files
        raise FindingsWorkbookError(str(exc)) from exc

    try:
        if FINDINGS_SHEET not in workbook.sheetnames:
            raise FindingsWorkbookError(
                f"Feuille « {FINDINGS_SHEET} » absente du fichier."
            )
        rows = list(workbook[FINDINGS_SHEET].iter_rows(values_only=True))
    finally:
        workbook.close()

    header_row, columns = _locate_columns(rows)

    data: list[tuple] = []
    pairs: list[tuple[str, str]] = []
    for row in rows[header_row:]:
        component = _cell(row, columns, "component")
        rule_id = _cell(row, columns, "rule")
        if not component and not rule_id:
            continue
        pairs.append((component, rule_id))
        data.append(row)

    return [
        ImportedFinding(
            key=key,
            qualification=FindingQualification.from_row(
                [_cell(row, columns, field) for field in _QUALIFICATION_FIELDS]
            ),
            finding=_finding(row, columns),
        )
        for key, row in zip(assign_keys(pairs), data)
    ]


def read_findings_qualifications(
    workbook_path: str | Path,
) -> dict[QualificationKey, FindingQualification]:
    """Filled-in TechLead columns of ``workbook_path``, keyed per finding."""
    return {
        row.key: row.qualification
        for row in read_findings_workbook(workbook_path)
        if not row.qualification.is_empty()
    }


def _locate_columns(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    """Index of the header row and the 1-based column of each field it names.

    The row index is 0-based so ``rows[header_row:]`` is the data; the
    column numbers are 1-based like everywhere else in the workbook code.
    """
    for index, row in enumerate(rows[:_HEADER_SEARCH_ROWS]):
        mapping: dict[str, int] = {}
        for column, value in enumerate(row, start=1):
            field = _FIELD_BY_HEADER.get(_normalise(value))
            if field is not None:
                mapping.setdefault(field, column)
        if len(mapping) >= _MIN_RECOGNISED_HEADERS:
            return index + 1, mapping
    return FIRST_DATA_ROW - 1, _DEFAULT_COLUMNS


def _normalise(value: object) -> str:
    """Header label stripped of everything that varies between files."""
    decomposed = unicodedata.normalize("NFKD", str(value if value is not None else ""))
    without_accents = "".join(
        char for char in decomposed if not unicodedata.combining(char)
    )
    return _NON_ALPHANUMERIC_RE.sub(" ", without_accents.lower()).strip()


def _cell(row: tuple, columns: dict[str, int], field: str) -> str:
    """Value of ``field`` in ``row``, tolerant to short rows and missing columns."""
    column = columns.get(field)
    if column is None or column > len(row):
        return ""
    value = row[column - 1]
    return "" if value is None else str(value).strip()


def _finding(row: tuple, columns: dict[str, int]) -> Finding:
    """Rebuild the finding a row describes, as far as the file allows.

    The rule is reconstructed from the exported columns only, so it carries
    no scope and no API version range. That is enough for the row to be
    listed, qualified and exported again; a rule Dewey still knows about is
    substituted by the caller.
    """
    category, _, subcategory = _cell(row, columns, "category").partition(" - ")
    message = _cell(row, columns, "message")
    rule = Rule(
        id=_cell(row, columns, "rule"),
        enabled=True,
        scope="",
        category=category.strip(),
        subcategory=subcategory.strip(),
        severity=_SEVERITY_CODES.get(_cell(row, columns, "severity"), "Info"),
        source=_cell(row, columns, "source"),
        reference=_cell(row, columns, "reference"),
        title=_cell(row, columns, "title"),
        description=message,
        rationale=_cell(row, columns, "rationale"),
        remediation=_cell(row, columns, "remediation"),
    )
    return Finding(
        rule=rule,
        target_kind=_cell(row, columns, "target_kind"),
        target_name=_cell(row, columns, "component"),
        message=message,
        details=[
            line.strip()
            for line in _cell(row, columns, "details").splitlines()
            if line.strip()
        ],
    )


_FIELD_BY_HEADER = {
    _normalise(header): field for header, field in zip(HEADERS, COLUMN_FIELDS)
}
_DEFAULT_COLUMNS = {field: column for column, field in enumerate(COLUMN_FIELDS, start=1)}
