"""Data Dictionary workbook generation for the Excel report writer.

Extracted from ``excel_writer.py`` to keep that module focused on the
smaller security/inventory/PMD workbooks. This mixin owns the logic that
splits large object catalogs across multiple ``.xlsx`` parts and renders
the per-object field sheets.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from src.core.models import ObjectInfo

# Excel spec allows a theoretically unbounded number of sheets per workbook but
# the file format becomes unresponsive well before that. We cap at a safe soft
# limit (1 index sheet + N object sheets) and roll over into a "part 2"
# workbook past the threshold.
MAX_OBJECT_SHEETS_PER_WORKBOOK = 200

# Excel forbids certain characters in sheet names and enforces a 31-char limit.
_FORBIDDEN_SHEET_CHARS_RE = re.compile(r"[:\\/?*\[\]]")


class _ExcelDataDictionaryMixin:
    """Adds Data Dictionary workbook generation to ``ExcelReportWriter``."""

    def write_data_dictionary_workbooks(
        self,
        objects: list[ObjectInfo],
        output_dir: str | Path,
        *,
        max_object_sheets: int = MAX_OBJECT_SHEETS_PER_WORKBOOK,
        filename_base: str = "data_dictionary",
        include_comment: bool = True,
        include_piloted_by: bool = True,
        include_status: bool = True,
        include_squad: bool = True,
        include_squad_consumer: bool = True,
        include_field_comment: bool = True,
        include_field_piloted_by: bool = True,
        concat_description: bool = True,
    ) -> list[Path]:
        """Generate the Data Dictionary workbook(s).

        Each workbook starts with a "Synthese" sheet listing the objects it
        contains (general info) followed by one sheet per object describing
        its fields. When the number of object sheets exceeds
        ``max_object_sheets`` a new workbook is created (``{filename_base}_part_2.xlsx``,
        ``..._part_3.xlsx`` and so on) so Excel stays responsive.

        ``include_comment``, ``include_piloted_by``, ``include_status``,
        ``include_squad`` and ``include_squad_consumer`` control whether the
        "Commentaire Dewey", "Piloté par", "Status", "Squad Responsable" and
        "Squad Consommatrice" columns (and matching per-object notes) are
        rendered at all. ``include_field_comment`` and
        ``include_field_piloted_by`` control whether the per-field
        "Commentaire Dewey" and "Piloté par" columns are rendered on each
        object's fields sheet. ``concat_description`` controls whether the
        "Commentaire Dewey" value concatenates the metadata Description or
        only shows the raw user-entered comment.

        Returns the list of written file paths in order.
        """
        output_base = Path(output_dir)
        output_base.mkdir(parents=True, exist_ok=True)

        # Skip objects that do not declare any field: their per-object sheet
        # would otherwise be empty (just the "Aucun champ detecte" hint) and
        # they pollute the Synthese sheet without bringing documentation value.
        documented_objects = [obj for obj in objects if obj.fields]
        skipped_count = len(objects) - len(documented_objects)
        if skipped_count > 0:
            self.log(
                f"Data Dictionary : {skipped_count} objet(s) sans champ ignore(s)."
            )

        if not documented_objects:
            # Still produce an (almost) empty workbook so that the index page
            # and the HTML preview pipeline expose the absence of data clearly.
            path = output_base / f"{filename_base}.xlsx"
            workbook = Workbook()
            summary = workbook.active
            summary.title = "Synthese"
            self._write_sheet(
                summary,
                self._data_dictionary_summary_headers(
                    include_comment,
                    include_piloted_by,
                    include_status,
                    include_squad,
                    include_squad_consumer,
                ),
                [],
            )
            workbook.save(path)
            self.log(f"Data Dictionary genere (aucun objet detecte) : {path}")
            return [path]

        ordered_objects = sorted(
            documented_objects, key=lambda obj: (obj.api_name or "").lower()
        )
        chunks: list[list[ObjectInfo]] = [
            ordered_objects[index : index + max_object_sheets]
            for index in range(0, len(ordered_objects), max_object_sheets)
        ]
        total_parts = len(chunks)
        written: list[Path] = []
        for part_index, chunk in enumerate(chunks, start=1):
            path = output_base / self._data_dictionary_filename(part_index, filename_base)
            self._write_data_dictionary_workbook(
                chunk,
                path,
                part_index=part_index,
                total_parts=total_parts,
                include_comment=include_comment,
                include_piloted_by=include_piloted_by,
                include_status=include_status,
                include_squad=include_squad,
                include_squad_consumer=include_squad_consumer,
                include_field_comment=include_field_comment,
                include_field_piloted_by=include_field_piloted_by,
                concat_description=concat_description,
            )
            written.append(path)
        summary = (
            f"Data Dictionary genere ({len(ordered_objects)} objets, "
            f"{total_parts} fichier(s)) : "
            + ", ".join(path.name for path in written)
        )
        self.log(summary)
        return written

    @staticmethod
    def _data_dictionary_filename(part_index: int, filename_base: str = "data_dictionary") -> str:
        if part_index <= 1:
            return f"{filename_base}.xlsx"
        return f"{filename_base}_part_{part_index}.xlsx"

    @staticmethod
    def _data_dictionary_summary_headers(
        include_comment: bool = True,
        include_piloted_by: bool = True,
        include_status: bool = True,
        include_squad: bool = True,
        include_squad_consumer: bool = True,
    ) -> list[str]:
        headers = [
            "API Name",
            "Label",
            "Label pluriel",
            "Custom",
            "Modele de partage",
            "Statut deploiement",
            "Visibilite",
            "Nb champs",
            "Nb champs custom",
            "Nb record types",
            "Nb validation rules",
            "Nb relations",
            "Feuille",
            "Description",
        ]
        if include_comment:
            headers.append("Commentaire Dewey")
        if include_piloted_by:
            headers.append("Piloté par")
        if include_status:
            headers.append("Status")
        if include_squad:
            headers.append("Squad Responsable")
        if include_squad_consumer:
            headers.append("Squad Consommatrice")
        return headers

    def _write_data_dictionary_workbook(
        self,
        objects_chunk: list[ObjectInfo],
        output_path: Path,
        *,
        part_index: int,
        total_parts: int,
        include_comment: bool = True,
        include_piloted_by: bool = True,
        include_status: bool = True,
        include_squad: bool = True,
        include_squad_consumer: bool = True,
        include_field_comment: bool = True,
        include_field_piloted_by: bool = True,
        concat_description: bool = True,
    ) -> None:
        workbook = Workbook()
        used_names: set[str] = set()
        # Reserve the summary sheet name up front so no object collides with it.
        summary_name = self._unique_sheet_name("Synthese", used_names)
        summary = workbook.active
        summary.title = summary_name

        sheet_names_by_object: list[tuple[ObjectInfo, str]] = []
        for obj in objects_chunk:
            sheet_name = self._unique_sheet_name(
                obj.api_name or "Objet", used_names
            )
            sheet_names_by_object.append((obj, sheet_name))

        summary_rows = []
        for obj, sheet_name in sheet_names_by_object:
            row = [
                obj.api_name,
                obj.label,
                obj.plural_label,
                "Oui" if obj.custom else "Non",
                obj.sharing_model,
                obj.deployment_status,
                obj.visibility,
                len(obj.fields),
                sum(1 for field in obj.fields if field.custom),
                len(obj.record_types),
                len(obj.validation_rules),
                len(obj.relationships),
                sheet_name,
                obj.description,
            ]
            if include_comment:
                row.append(obj.dewey_comment_combined if concat_description else (obj.dewey_comment or ""))
            if include_piloted_by:
                row.append(obj.dewey_piloted_by)
            if include_status:
                row.append(obj.dewey_status)
            if include_squad:
                row.append(obj.dewey_squad)
            if include_squad_consumer:
                row.append(obj.dewey_squad_consumer)
            summary_rows.append(row)
        self._write_sheet(
            summary,
            self._data_dictionary_summary_headers(
                include_comment,
                include_piloted_by,
                include_status,
                include_squad,
                include_squad_consumer,
            ),
            summary_rows,
        )

        if total_parts > 1:
            # Add a small indicator on the summary sheet so the user knows
            # other parts exist without having to open a file explorer.
            note_row = summary.max_row + 2
            summary.cell(
                row=note_row,
                column=1,
                value=(
                    f"Partie {part_index} / {total_parts}. "
                    "Les objets suivants se trouvent dans les autres fichiers "
                    "data_dictionary_part_*.xlsx."
                ),
            ).font = Font(italic=True)

        for obj, sheet_name in sheet_names_by_object:
            worksheet = workbook.create_sheet(sheet_name)
            self._write_object_fields_sheet(
                worksheet,
                obj,
                include_comment=include_comment,
                include_piloted_by=include_piloted_by,
                include_status=include_status,
                include_squad=include_squad,
                include_squad_consumer=include_squad_consumer,
                include_field_comment=include_field_comment,
                include_field_piloted_by=include_field_piloted_by,
                concat_description=concat_description,
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _write_object_fields_sheet(
        self,
        worksheet,
        obj: ObjectInfo,
        *,
        include_comment: bool = True,
        include_piloted_by: bool = True,
        include_status: bool = True,
        include_squad: bool = True,
        include_squad_consumer: bool = True,
        include_field_comment: bool = True,
        include_field_piloted_by: bool = True,
        concat_description: bool = True,
    ) -> None:
        headers = [
            "API Name",
            "Label",
            "Type",
            "Obligatoire",
            "Custom",
            "Reference vers",
            "Relationship Name",
            "Description",
        ]
        # Dewey-authored per-field columns, appended after the raw metadata
        # columns, each independently toggled from the UI.
        if include_field_comment:
            headers.append("Commentaire Dewey")
        if include_field_piloted_by:
            headers.append("Piloté par")

        rows = []
        for field in obj.fields:
            row = [
                field.api_name,
                field.label,
                field.data_type,
                "Oui" if field.required else "Non",
                "Oui" if field.custom else "Non",
                ", ".join(field.reference_to),
                field.relationship_name,
                field.description,
            ]
            if include_field_comment:
                row.append(field.dewey_comment)
            if include_field_piloted_by:
                row.append(field.dewey_piloted_by)
            rows.append(row)
        self._write_sheet(worksheet, headers, rows)

        if not rows:
            # Leave a tiny hint explaining why the sheet is empty rather than
            # letting the user wonder if parsing failed.
            worksheet.cell(
                row=2,
                column=1,
                value="Aucun champ detecte dans la metadata pour cet objet.",
            ).font = Font(italic=True, color="666666")

        note_row = worksheet.max_row + 1
        comment_value = obj.dewey_comment_combined if concat_description else (obj.dewey_comment or "")
        if include_comment and comment_value:
            note_row += 1
            worksheet.cell(
                row=note_row,
                column=1,
                value=f"Commentaire Dewey : {comment_value}",
            ).font = Font(italic=True, bold=True)
        if include_piloted_by and obj.dewey_piloted_by:
            note_row += 1
            worksheet.cell(
                row=note_row,
                column=1,
                value=f"Piloté par : {obj.dewey_piloted_by}",
            ).font = Font(italic=True, bold=True)
        if include_status and obj.dewey_status and obj.dewey_status != "-":
            note_row += 1
            worksheet.cell(
                row=note_row,
                column=1,
                value=f"Status : {obj.dewey_status}",
            ).font = Font(italic=True, bold=True)
        if include_squad and obj.dewey_squad:
            note_row += 1
            worksheet.cell(
                row=note_row,
                column=1,
                value=f"Squad Responsable : {obj.dewey_squad}",
            ).font = Font(italic=True, bold=True)
        if include_squad_consumer and obj.dewey_squad_consumer:
            note_row += 1
            worksheet.cell(
                row=note_row,
                column=1,
                value=f"Squad Consommatrice : {obj.dewey_squad_consumer}",
            ).font = Font(italic=True, bold=True)

    @staticmethod
    def _unique_sheet_name(desired: str, used: set[str]) -> str:
        """Return a unique, Excel-compliant sheet name and register it."""
        cleaned = _FORBIDDEN_SHEET_CHARS_RE.sub("_", desired or "").strip()
        cleaned = cleaned.strip("'")  # Excel rejects names wrapped in quotes
        if not cleaned:
            cleaned = "Feuille"
        base = cleaned[:31]
        candidate = base
        counter = 1
        # Case-insensitive comparison (Excel treats sheet names this way).
        existing_lower = {name.lower() for name in used}
        while candidate.lower() in existing_lower:
            counter += 1
            suffix = f"~{counter}"
            truncated = base[: max(1, 31 - len(suffix))]
            candidate = f"{truncated}{suffix}"
        used.add(candidate)
        return candidate
