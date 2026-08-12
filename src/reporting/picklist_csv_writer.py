"""Export picklist values to individual CSV files, plus a summary workbook.

Wired to the "Documentation" menu ("Creer les CSV des picklists"). Unlike
:meth:`ExcelReportWriter.write_picklists_workbook` (a single inventory sheet
listing every picklist), this produces one CSV per picklist field so each can
be reused independently (e.g. imported into another tool, a data loader
template, or a translation workbook).

Layout created under ``output_dir``::

    picklist/
      global/<GlobalValueSetName>.csv
      fields/<Object>_<Field>.csv
      picklists_summary.xlsx
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.models import FieldInfo, ObjectInfo

LogCallback = Callable[[str], None]

_INVALID_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')

_SUMMARY_HEADERS = [
    "Object",
    "Champs",
    "Nom de la Global Picklist",
    "Repertoire fichier",
    "Nom Fichier",
]


def _safe_filename(name: str) -> str:
    cleaned = _INVALID_FILENAME_CHARS_RE.sub("_", (name or "").strip())
    return cleaned or "picklist"


class PicklistCsvWriter:
    """Writes one CSV per picklist field (global or local) and a summary workbook."""

    def __init__(self, log_callback: LogCallback | None = None) -> None:
        self.log: LogCallback = log_callback or (lambda message: None)

    def write_picklist_csv_export(
        self, objects: list[ObjectInfo], output_dir: str | Path
    ) -> Path:
        """Create the ``picklist/`` folder tree and return the summary workbook path."""
        root = Path(output_dir) / "picklist"
        fields_dir = root / "fields"
        global_dir = root / "global"
        fields_dir.mkdir(parents=True, exist_ok=True)
        global_dir.mkdir(parents=True, exist_ok=True)

        summary_rows: list[list[object]] = []
        written_global_files: set[str] = set()
        field_csv_count = 0

        for obj in objects:
            for item in obj.fields:
                if not item.is_picklist:
                    continue

                if item.picklist_is_global:
                    directory_label = "global"
                    filename = f"{_safe_filename(item.picklist_global_name or item.api_name)}.csv"
                    if filename not in written_global_files:
                        self._write_picklist_values_csv(global_dir / filename, item)
                        written_global_files.add(filename)
                else:
                    directory_label = "fields"
                    filename = f"{_safe_filename(obj.api_name)}_{_safe_filename(item.api_name)}.csv"
                    self._write_picklist_values_csv(fields_dir / filename, item)
                    field_csv_count += 1

                summary_rows.append(
                    [
                        obj.api_name,
                        item.api_name,
                        item.picklist_global_name if item.picklist_is_global else "-",
                        directory_label,
                        filename,
                    ]
                )

        summary_path = root / "picklists_summary.xlsx"
        self._write_summary_workbook(summary_path, summary_rows)

        self.log(
            "Export CSV des picklists termine : "
            f"{len(summary_rows)} champ(s) picklist au total, "
            f"{field_csv_count} fichier(s) dans fields/, "
            f"{len(written_global_files)} fichier(s) dans global/ : {root}"
        )
        return summary_path

    @staticmethod
    def _write_picklist_values_csv(path: Path, item: FieldInfo) -> None:
        labels = item.picklist_values
        api_names = item.picklist_api_names
        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow(["Label", "Nom API"])
            for index, label in enumerate(labels):
                api_name = api_names[index] if index < len(api_names) else label
                writer.writerow([label, api_name])

    def _write_summary_workbook(self, path: Path, rows: list[list[object]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Picklists"
        sheet.append(_SUMMARY_HEADERS)
        header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in rows:
            sheet.append(row)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, header in enumerate(_SUMMARY_HEADERS, start=1):
            max_length = len(header)
            for row_cells in sheet.iter_rows(min_col=index, max_col=index, min_row=2):
                value = row_cells[0].value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 60)

        workbook.save(path)
