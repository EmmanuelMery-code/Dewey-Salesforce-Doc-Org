"""Export the analyzer findings to the "document des findings" workbook.

Wired to the "Documentation" menu ("Creer le document des findings"). The
layout mirrors the qualification workbook used by the TechLeads: a
``Findings`` sheet with three column groups (finding / qualification / US)
and a static ``Legende`` sheet describing the statuses and the workflow.

Dewey owns the ``A..L`` columns. The qualification (``M..P``) and US
(``Q..S``) columns belong to the TechLead reviewing the run: they are styled
and carry the status dropdown, and are filled from the values previously
imported back from a reviewed workbook (see
:mod:`src.core.findings_qualification`), or left empty when there are none.

The findings a run no longer reports are exported too — the caller passes
them along with ``resolved_keys`` — with their status forced to "Terminé".
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Collection, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.analyzer.models import Finding
from src.core.findings_qualification import (
    RESOLVED_STATUS,
    FindingQualification,
    QualificationKey,
    finding_keys,
    sort_findings,
)

LogCallback = Callable[[str], None]

FINDINGS_SHEET = "Findings"
LEGEND_SHEET = "Légende"

# Row 1 groups the columns; row 2 holds the actual headers.
_GROUP_BASE = 12  # A..L — the finding as detected by Dewey
_GROUP_QUALIFICATION = 4  # M..P — the TechLead's qualification
_GROUP_US = 3  # Q..S — the TechLead's US description

#: First row carrying a finding, used as a fallback by the importer when it
#: cannot recognise the header row of the file it is given.
FIRST_DATA_ROW = 3

HEADERS = [
    "Sévérité",
    "Catégorie",
    "Règle",
    "Type composant",
    "Composant",
    "Titre du finding",
    "Message",
    "Justification",
    "Remédiation",
    "Source",
    "Référence",
    "Détails",
    "Statut",
    "Équipe",
    "Sprint cible",
    "Numéro US",
    "Titre US",
    "Description US",
    "Critères d'acceptation",
]

#: Field fed by each column, aligned with :data:`HEADERS`. The importer maps
#: a header label back onto its field through it, which is what lets a
#: workbook whose columns were shifted still be read correctly.
COLUMN_FIELDS = (
    "severity",
    "category",
    "rule",
    "target_kind",
    "component",
    "title",
    "message",
    "rationale",
    "remediation",
    "source",
    "reference",
    "details",
    "status",
    "team",
    "target_sprint",
    "us_number",
    "us_title",
    "us_description",
    "acceptance_criteria",
)

_COLUMN_WIDTHS = [12, 22, 14, 16, 26, 34, 38, 45, 45, 28, 40, 40, 14, 18, 24, 33.5, 53.5, 52, 52]

# Colours carry the explicit "FF" alpha channel so the workbook matches the
# reference file produced by Excel itself.
_DARK = "FF2D3748"
_BLUE = "FF4A6FA5"
_RED = "FFC0000C"
_BORDER_COLOR = "FFD1D5DB"
_WHITE = "FFFFFFFF"
_BLACK = "FF000000"
_BASE_FILL = "FFFFFFFF"
_STATUS_FILL = "FFFEF9C3"
_QUALIFICATION_FILL = "FFEEF2FA"
_US_FILL = "FFFEF0EE"
_LEGEND_FILL = "FFF8FAFC"

SEVERITY_LABELS = {
    "Critical": "Critique",
    "Major": "Majeur",
    "Minor": "Mineur",
    "Info": "Info",
}
_SEVERITY_COLORS = {
    "Critique": "FFC0392B",
    "Majeur": "FFE67E22",
    "Mineur": "FFF9D057",
    "Info": "FF5B8DB8",
}

#: Allowed values of the "Statut" column, shared with the findings screen so
#: its editor offers exactly what the workbook's dropdown accepts.
STATUSES = ["À traiter", "Faux positif", "En cours", RESOLVED_STATUS]

_LEGEND_STATUSES = [
    ("À traiter", "Valeur par défaut — finding à qualifier par le TechLead"),
    ("Faux positif", "Finding non pertinent — sera exclu des prochaines analyses Dewey"),
    ("En cours", "Remédiation en cours dans un sprint"),
    (
        RESOLVED_STATUS,
        "Remédiation validée, finding résolu — position mise d'office par Dewey "
        "sur les findings que l'analyse ne détecte plus",
    ),
]
_LEGEND_WORKFLOW = [
    (
        "Faux positif",
        "Remplir Statut = Faux positif. Remonter à l'équipe Dewey pour exclusion de la règle.",
    ),
    (
        "Vrai sujet",
        "Remplir Statut, Équipe, Sprint cible. Copier les 3 colonnes US dans Jira pour créer l'US.",
    ),
]

_HEADER_ROW_HEIGHT = 32.15
_TITLE_ROW_HEIGHT = 22.0
_DATA_ROW_HEIGHT = 75.0
_LEGEND_ROW_HEIGHT = 28.0

_BORDER = Border(*(Side(style="thin", color=_BORDER_COLOR) for _ in range(4)))

_INVALID_FILENAME_CHARS_RE = re.compile(r"[^A-Za-z0-9_-]+")


def findings_workbook_path(
    excel_dir: str | Path, alias: str = "", exported_at: datetime | None = None
) -> Path:
    """Destination of the findings workbook for ``alias`` inside ``excel_dir``.

    The name carries the export time down to the second, so regenerating the
    documentation never overwrites a workbook a TechLead may still be filling
    in. Note this is the moment of the export, not the date of the analysed
    run, which the sheet title carries instead.
    """
    stamp = (exported_at or datetime.now()).strftime("%Y%m%d_%H%M%S")
    slug = _INVALID_FILENAME_CHARS_RE.sub("_", alias or "").strip("_") or "org"
    return Path(excel_dir) / f"Dewey_Findings_{slug}_{stamp}.xlsx"


def severity_label(finding: Finding) -> str:
    """French severity label used by the workbook and the findings screen."""
    return SEVERITY_LABELS.get(finding.rule.severity, finding.rule.severity or "Info")


def _category_label(finding: Finding) -> str:
    rule = finding.rule
    if rule.subcategory:
        return f"{rule.category} - {rule.subcategory}"
    return rule.category


class FindingsExcelWriter:
    """Writes the two-sheet findings qualification workbook."""

    def __init__(self, log_callback: LogCallback | None = None) -> None:
        self.log: LogCallback = log_callback or (lambda message: None)

    def write_findings_workbook(
        self,
        findings: Sequence[Finding],
        output_path: str | Path,
        *,
        alias: str = "",
        run_date: date | None = None,
        qualifications: Mapping[QualificationKey, FindingQualification] | None = None,
        resolved_keys: Collection[QualificationKey] = (),
    ) -> Path:
        """Write ``findings`` to ``output_path`` and return the written path.

        ``qualifications`` pre-fills the TechLead columns from a previously
        imported workbook; findings without a stored entry keep them empty.
        ``resolved_keys`` designates the findings the analyzer no longer
        reports, whose status is forced to "Terminé".
        """
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        resolved = set(resolved_keys)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = FINDINGS_SHEET
        self._write_findings_sheet(
            sheet,
            findings,
            alias=alias,
            run_date=run_date,
            qualifications=qualifications or {},
            resolved=resolved,
        )
        self._write_legend_sheet(workbook.create_sheet(LEGEND_SHEET))
        workbook.save(output)

        closed = f", dont {len(resolved)} resolu(s)" if resolved else ""
        self.log(
            f"Document des findings genere : {len(findings)} finding(s)"
            f"{closed} - {output}"
        )
        return output

    # ------------------------------------------------------------------ findings

    def _write_findings_sheet(
        self,
        sheet,
        findings: Sequence[Finding],
        *,
        alias: str,
        run_date: date | None,
        qualifications: Mapping[QualificationKey, FindingQualification],
        resolved: set[QualificationKey],
    ) -> None:
        stamp = (run_date or date.today()).strftime("%d/%m/%Y")
        title = f"Finding ({f'{alias} {stamp}'.strip()})"
        self._write_group_titles(sheet, title)
        self._write_headers(sheet)

        ordered = sort_findings(findings)
        keys = finding_keys(ordered)
        for offset, (finding, key) in enumerate(zip(ordered, keys)):
            self._write_finding_row(
                sheet,
                FIRST_DATA_ROW + offset,
                finding,
                qualifications.get(key),
                resolved=key in resolved,
            )

        for column, width in enumerate(_COLUMN_WIDTHS, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A3"

        if findings:
            last_row = 2 + len(findings)
            validation = DataValidation(
                type="list",
                formula1='"' + ",".join(STATUSES) + '"',
                allow_blank=False,
            )
            sheet.add_data_validation(validation)
            validation.add(f"M3:M{last_row}")

    def _write_group_titles(self, sheet, title: str) -> None:
        groups = [
            (title, _DARK, _GROUP_BASE),
            ("Qualification", _BLUE, _GROUP_QUALIFICATION),
            ("US", _RED, _GROUP_US),
        ]
        column = 1
        for label, color, span in groups:
            cell = sheet.cell(row=1, column=column, value=label)
            cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for offset in range(span):
                sheet.cell(row=1, column=column + offset).border = _BORDER
            sheet.merge_cells(
                start_row=1,
                start_column=column,
                end_row=1,
                end_column=column + span - 1,
            )
            column += span
        sheet.row_dimensions[1].height = _TITLE_ROW_HEIGHT

    def _write_headers(self, sheet) -> None:
        for column, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=2, column=column, value=header)
            cell.font = Font(name="Arial", size=9, bold=True, color=_WHITE)
            cell.fill = PatternFill(fill_type="solid", fgColor=self._group_color(column))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
        sheet.row_dimensions[2].height = _HEADER_ROW_HEIGHT

    def _write_finding_row(
        self,
        sheet,
        row: int,
        finding: Finding,
        qualification: FindingQualification | None = None,
        *,
        resolved: bool = False,
    ) -> None:
        rule = finding.rule
        severity = severity_label(finding)
        values: list[object | None] = [
            severity,
            _category_label(finding),
            rule.id,
            finding.target_kind,
            finding.target_name,
            rule.title,
            finding.message or rule.description,
            rule.rationale,
            rule.remediation,
            rule.source,
            rule.reference,
            "\n".join(finding.details) or None,
        ]
        # Qualification (M..P) and US (Q..S) are the TechLead's own columns:
        # restored from a previously imported workbook, or left empty. The
        # status is the one exception Dewey writes by itself, on the findings
        # it no longer detects.
        techlead = qualification or FindingQualification()
        if resolved:
            techlead = techlead.with_status(RESOLVED_STATUS)
        values += [value or None for value in techlead.as_row()]

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _BORDER
            if column == 1:
                cell.font = Font(name="Arial", size=9, bold=True, color=_WHITE)
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=_SEVERITY_COLORS.get(severity, _SEVERITY_COLORS["Info"]),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill(fill_type="solid", fgColor=self._body_fill(column))
            horizontal = "center" if column == 13 else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)

        sheet.row_dimensions[row].height = _DATA_ROW_HEIGHT

    @staticmethod
    def _group_color(column: int) -> str:
        if column <= _GROUP_BASE:
            return _DARK
        if column <= _GROUP_BASE + _GROUP_QUALIFICATION:
            return _BLUE
        return _RED

    @staticmethod
    def _body_fill(column: int) -> str:
        if column <= _GROUP_BASE:
            return _BASE_FILL
        if column == 13:
            return _STATUS_FILL
        if column <= _GROUP_BASE + _GROUP_QUALIFICATION:
            return _QUALIFICATION_FILL
        return _US_FILL

    # ------------------------------------------------------------------ legend

    def _write_legend_sheet(self, sheet) -> None:
        # (label, description, section header colour, label fill)
        rows: list[tuple[str | None, str | None, str | None, str]] = [
            (None, None, None, _BASE_FILL),
            ("STATUTS", None, _BLUE, _BLUE),
        ]
        rows.extend(
            (label, description, None, _QUALIFICATION_FILL)
            for label, description in _LEGEND_STATUSES
        )
        rows.append((None, None, None, _BASE_FILL))
        rows.append(("WORKFLOW", None, _DARK, _DARK))
        rows.extend(
            (label, description, None, _LEGEND_FILL)
            for label, description in _LEGEND_WORKFLOW
        )

        for index, (label, description, section_color, label_fill) in enumerate(rows, start=1):
            left = sheet.cell(row=index, column=1, value=label)
            right = sheet.cell(row=index, column=2, value=description)
            for cell in (left, right):
                cell.border = _BORDER
            left.font = Font(name="Arial", size=9, color=_BLACK)
            left.alignment = Alignment(vertical="center")
            left.fill = PatternFill(fill_type="solid", fgColor=label_fill)
            right.font = Font(name="Arial", size=9)
            right.alignment = Alignment(vertical="center", wrap_text=True)
            right.fill = PatternFill(fill_type="solid", fgColor=_LEGEND_FILL)

            if section_color is not None:
                left.font = Font(name="Arial", size=9, bold=True, color=_WHITE)
                sheet.merge_cells(start_row=index, start_column=1, end_row=index, end_column=2)
            sheet.row_dimensions[index].height = _LEGEND_ROW_HEIGHT

        sheet.column_dimensions["A"].width = 26
        sheet.column_dimensions["B"].width = 70
