"""Classeur Excel equivalent au sous-onglet HTML "PSet Group Summary".

Chaque onglet reprend une section des deux pages HTML : les cadrans et la
matrice du sous-onglet, puis les tableaux de la page de detail
``psg_summary_details.html``. Les chiffres viennent du meme calcul
(:mod:`src.core.psg_access`), donc les deux formats disent la meme chose.
"""

from __future__ import annotations

from pathlib import Path
from typing import Collection

from openpyxl import Workbook

from src.core.models import MetadataSnapshot
from src.core.psg_access import (
    ALL_FLAGS,
    COVERAGE_REASONS,
    CRUD_FLAGS,
    SHARING_FLAGS,
    STATUS_HELP,
    GroupAccess,
    build_group_access,
    covered_object_names,
    listed_object_names,
    sharing_context,
)

YES = "Oui"
NO = ""


def _codes(access, flags: tuple[tuple[str, str, str], ...]) -> str:
    """Codes des droits accordes, comme les badges allumes de la matrice HTML."""

    if access is None:
        return ""
    return " ".join(code for code, _label, attribute in flags if access.granted(attribute))


def _group_title(access: GroupAccess) -> str:
    return access.group.label or access.group.name


class _ExcelPsgSummaryMixin:
    """Ajoute ``write_psg_summary_workbook`` a :class:`ExcelReportWriter`."""

    def write_psg_summary_workbook(
        self,
        snapshot: MetadataSnapshot,
        output_path: str | Path,
        selected_objects: Collection[str] | None = None,
    ) -> Path:
        """Ecrit le classeur de synthese des Permission Set Groups.

        ``selected_objects`` reprend la selection de l'ecran Data Dictionary :
        elle devient une colonne filtrable, equivalente a la case a cocher
        "filtrer pour les objets selectionnes" de la page HTML.
        """

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        accesses = build_group_access(snapshot)
        owd, rule_counts = sharing_context(snapshot)
        covered = covered_object_names(accesses)
        object_names = listed_object_names(snapshot, accesses)
        selected = {name for name in (selected_objects or ()) if name}

        workbook = Workbook()
        self._write_sheet(
            workbook.active,
            ["Indicateur", "Valeur", "Precision", "Onglet de detail"],
            self._psg_kpi_rows(accesses, covered, object_names),
        )
        workbook.active.title = "Synthese"

        self._write_sheet(
            workbook.create_sheet("Matrice"),
            self._psg_matrix_headers(accesses, selected),
            self._psg_matrix_rows(accesses, object_names, owd, rule_counts, selected),
        )
        self._write_sheet(
            workbook.create_sheet("DroitsParGroupe"),
            [
                "Groupe",
                "Nom API du groupe",
                "Statut",
                "Objet",
                "OWD",
                *[label for _code, label, _attribute in ALL_FLAGS],
                "Accorde par",
            ],
            self._psg_group_rows(accesses, owd),
        )
        self._write_sheet(
            workbook.create_sheet("Groupes"),
            [
                "Nom API",
                "Label",
                "Statut",
                "PS membres",
                "PS non analyses",
                "Objets couverts",
                "Objets avec Modify All",
                "Objets avec View All",
                "Description",
            ],
            self._psg_groups_rows(accesses),
        )
        self._write_sheet(
            workbook.create_sheet("PermissionSets"),
            ["Permission Set", "Etat", "Nb groupes", "Groupes", "Droits objet"],
            self._psg_permission_set_rows(snapshot, accesses),
        )
        self._write_sheet(
            workbook.create_sheet("Couverture"),
            ["Objet", "OWD", "Objet analyse", "Couvert", "Nb groupes", "Groupes"],
            self._psg_coverage_rows(accesses, object_names, owd),
        )
        for sheet_name, attribute in (
            ("ModifyAll", "modify_all_records"),
            ("ViewAll", "view_all_records"),
        ):
            self._write_sheet(
                workbook.create_sheet(sheet_name),
                ["Groupe", "Statut", "Objet", "Accorde par"],
                self._psg_wide_access_rows(accesses, attribute),
            )
        self._write_sheet(
            workbook.create_sheet("Legende"),
            ["Rubrique", "Valeur", "Signification"],
            self._psg_legend_rows(),
        )

        workbook.save(output)
        self.log(f"Classeur PSet Group Summary genere: {output}")
        return output

    def _psg_kpi_rows(
        self,
        accesses: list[GroupAccess],
        covered: set[str],
        object_names: list[str],
    ) -> list[list[object]]:
        active = sum(
            1 for item in accesses if (item.group.status or "").lower() == "updated"
        )
        members = {name for item in accesses for name in item.group.permission_sets if name}
        unresolved = {name for item in accesses for name in item.unresolved_permission_sets}
        modify_all = len(self._psg_wide_access_rows(accesses, "modify_all_records"))
        view_all = len(self._psg_wide_access_rows(accesses, "view_all_records"))
        return [
            [
                "Permission Set Groups",
                len(accesses),
                f"{active} au statut Updated",
                "Groupes",
            ],
            [
                "Permission Sets membres",
                len(members),
                f"{len(unresolved)} non analyse(s)",
                "PermissionSets",
            ],
            [
                "Objets couverts",
                len(covered),
                f"sur {len(object_names)} objet(s) listes",
                "Couverture",
            ],
            [
                "Couples groupe/objet avec Modify All",
                modify_all,
                "Droit le plus large : contourne le partage",
                "ModifyAll",
            ],
            [
                "Couples groupe/objet avec View All",
                view_all,
                "Lecture de tous les enregistrements",
                "ViewAll",
            ],
        ]

    def _psg_matrix_headers(
        self,
        accesses: list[GroupAccess],
        selected: set[str],
    ) -> list[str]:
        headers = ["Objet", "OWD", "Regles de partage", "Couvert"]
        if selected:
            headers.append("Selectionne (Data Dictionary)")
        for access in accesses:
            title = _group_title(access)
            headers.extend([f"{title} - CRUD", f"{title} - Sharing & Visibility"])
        return headers

    def _psg_matrix_rows(
        self,
        accesses: list[GroupAccess],
        object_names: list[str],
        owd: dict[str, str],
        rule_counts: dict[str, int],
        selected: set[str],
    ) -> list[list[object]]:
        rows: list[list[object]] = []
        for object_name in object_names:
            entries = [access.objects.get(object_name) for access in accesses]
            row: list[object] = [
                object_name,
                owd.get(object_name, ""),
                rule_counts.get(object_name, 0),
                YES if any(entry is not None for entry in entries) else NO,
            ]
            if selected:
                row.append(YES if object_name in selected else NO)
            for entry in entries:
                row.extend([_codes(entry, CRUD_FLAGS), _codes(entry, SHARING_FLAGS)])
            rows.append(row)
        return rows

    def _psg_group_rows(
        self,
        accesses: list[GroupAccess],
        owd: dict[str, str],
    ) -> list[list[object]]:
        rows: list[list[object]] = []
        for access in accesses:
            for object_name in sorted(access.objects, key=str.lower):
                entry = access.objects[object_name]
                rows.append(
                    [
                        _group_title(access),
                        access.group.name,
                        access.group.status,
                        object_name,
                        owd.get(object_name, ""),
                        *[
                            YES if entry.granted(attribute) else NO
                            for _code, _label, attribute in ALL_FLAGS
                        ],
                        ", ".join(entry.contributors),
                    ]
                )
        return rows

    def _psg_groups_rows(self, accesses: list[GroupAccess]) -> list[list[object]]:
        return [
            [
                access.group.name,
                access.group.label,
                access.group.status,
                len(access.group.permission_sets),
                ", ".join(access.unresolved_permission_sets),
                len(access.objects),
                sum(
                    1
                    for entry in access.objects.values()
                    if entry.granted("modify_all_records")
                ),
                sum(
                    1
                    for entry in access.objects.values()
                    if entry.granted("view_all_records")
                ),
                access.group.description,
            ]
            for access in accesses
        ]

    def _psg_permission_set_rows(
        self,
        snapshot: MetadataSnapshot,
        accesses: list[GroupAccess],
    ) -> list[list[object]]:
        permission_sets = {item.name: item for item in snapshot.permission_sets}
        groups_by_member: dict[str, list[str]] = {}
        for access in accesses:
            for member in access.group.permission_sets:
                if member:
                    groups_by_member.setdefault(member, []).append(access.group.name)

        rows: list[list[object]] = []
        for member in sorted(groups_by_member, key=str.lower):
            artifact = permission_sets.get(member)
            groups = sorted(groups_by_member[member], key=str.lower)
            rows.append(
                [
                    member,
                    "Analyse" if artifact is not None else "Non analyse",
                    len(groups),
                    ", ".join(groups),
                    len(artifact.object_permissions) if artifact is not None else "",
                ]
            )
        return rows

    def _psg_coverage_rows(
        self,
        accesses: list[GroupAccess],
        object_names: list[str],
        owd: dict[str, str],
    ) -> list[list[object]]:
        groups_by_object: dict[str, list[str]] = {}
        for access in accesses:
            for object_name in access.objects:
                groups_by_object.setdefault(object_name, []).append(access.group.name)

        rows: list[list[object]] = []
        for object_name in object_names:
            groups = sorted(groups_by_object.get(object_name, []), key=str.lower)
            rows.append(
                [
                    object_name,
                    owd.get(object_name, ""),
                    YES if object_name in owd else NO,
                    YES if groups else NO,
                    len(groups),
                    ", ".join(groups),
                ]
            )
        return rows

    def _psg_wide_access_rows(
        self,
        accesses: list[GroupAccess],
        attribute: str,
    ) -> list[list[object]]:
        rows: list[list[object]] = []
        for access in accesses:
            for object_name in sorted(access.objects, key=str.lower):
                entry = access.objects[object_name]
                if not entry.granted(attribute):
                    continue
                rows.append(
                    [
                        _group_title(access),
                        access.group.status,
                        object_name,
                        ", ".join(entry.sources(attribute)),
                    ]
                )
        return rows

    def _psg_legend_rows(self) -> list[list[object]]:
        rows: list[list[object]] = [
            ["CRUD", code, label] for code, label, _attribute in CRUD_FLAGS
        ]
        rows.extend(
            ["Sharing & Visibility", code, label]
            for code, label, _attribute in SHARING_FLAGS
        )
        rows.extend(["Statut du groupe", value, description] for value, description in STATUS_HELP)
        rows.extend(
            ["Couverture d'un objet", title, explanation]
            for title, explanation in COVERAGE_REASONS
        )
        rows.append(
            [
                "Principe de calcul",
                "Union des permission sets membres",
                "Un Permission Set Group n'accorde aucun droit par lui-meme : les "
                "droits presentes sont l'union des objectPermissions de ses "
                "permission sets membres.",
            ]
        )
        return rows
