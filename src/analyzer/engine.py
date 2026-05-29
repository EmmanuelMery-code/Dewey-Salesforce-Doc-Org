from __future__ import annotations

import re
from pathlib import Path

from src.analyzer.apex_analyzer import (
    _strip_comments_and_strings,
    analyze_apex_artifact,
)
from src.analyzer.flow_analyzer import analyze_flow
from src.analyzer.models import Finding, Rule, SEVERITY_ORDER
from src.analyzer.object_analyzer import analyze_object, analyze_validation_rule
from src.analyzer.omni_analyzer import analyze_data_transform
from src.analyzer.rule_catalog import RuleCatalog
from src.core.models import (
    AgentInfo,
    ApexArtifact,
    FlowInfo,
    GenAiPromptInfo,
    MetadataSnapshot,
    ObjectInfo,
    ValidationRuleInfo,
)


IDENTIFIER_RE = re.compile(r"\b([A-Za-z_][A-Za-z0-9_]*)\b")


class AnalyzerEngine:
    """Orchestrateur de l'analyse statique ; retourne un ensemble de findings par artefact."""

    def __init__(self, catalog: RuleCatalog | None = None, exclusion_path: Path | str | None = None) -> None:
        self.catalog = catalog or RuleCatalog.load()
        self.rule_exclusions: dict[str, set[str]] = {}  # rule_id -> set of metadata names
        
        if exclusion_path:
            self.exclusion_path = Path(exclusion_path)
        else:
            # On cherche exclusion_PV.xlsx ou exclusion.xlsx
            app_root = Path(__file__).resolve().parent.parent.parent
            candidate_pv = app_root / "exclusion_PV.xlsx"
            candidate_std = app_root / "exclusion.xlsx"
            if candidate_pv.exists():
                self.exclusion_path = candidate_pv
            elif candidate_std.exists():
                self.exclusion_path = candidate_std
            else:
                self.exclusion_path = None
        
        if self.exclusion_path:
            self._load_rule_exclusions()

    def _load_rule_exclusions(self) -> None:
        """Charge les exclusions de règles spécifiques par métadonnée depuis l'Excel.
        
        Format attendu dans l'onglet 'exclusions regles' :
        Colonne A : Type Metadata (optionnel, pour lisibilité)
        Colonne B : Nom Metadata
        Colonne C : ID Règle (ou 'all')
        """
        if not self.exclusion_path or not self.exclusion_path.exists():
            return

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(self.exclusion_path, data_only=True, read_only=True)
            
            sheet = None
            for name in workbook.sheetnames:
                if name.strip().lower() in ("exclusions regles", "exclusions règles"):
                    sheet = workbook[name]
                    break
            
            if sheet:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 3:
                        continue
                    
                    metadata_name = str(row[1] or "").strip()
                    rule_id = str(row[2] or "").strip()
                    
                    if metadata_name and rule_id:
                        if rule_id.lower() == "all":
                            # On pourrait gérer 'all' ici si besoin, mais le parser global le fait déjà.
                            # Ici on se concentre sur les exclusions de règles spécifiques.
                            pass
                        
                        self.rule_exclusions.setdefault(rule_id, set()).add(metadata_name.lower())
            
            workbook.close()
        except Exception:
            # On ignore silencieusement les erreurs de lecture Excel pour ne pas bloquer l'analyse
            pass

    def _is_rule_applicable(self, rule: Rule, metadata_name: str, api_version: str | None = None) -> bool:
        """Vérifie si une règle doit être appliquée à une métadonnée donnée."""
        # 1. Vérification de l'exclusion spécifique
        if rule.id in self.rule_exclusions:
            if metadata_name.lower() in self.rule_exclusions[rule.id]:
                return False
        
        # 2. Vérification de la version d'API
        if api_version:
            try:
                version = float(api_version)
                if rule.min_api_version is not None and version < rule.min_api_version:
                    return False
                if rule.max_api_version is not None and version > rule.max_api_version:
                    return False
            except (ValueError, TypeError):
                pass
                
        return True

    # ------------------------------------------------------------------ per-artifact API

    def analyze_apex(self, artifact: ApexArtifact) -> list[Finding]:
        findings = analyze_apex_artifact(artifact, self.catalog)
        filtered = [f for f in findings if self._is_rule_applicable(f.rule, artifact.name, artifact.api_version)]
        return _sorted(filtered)

    def analyze_flow(self, flow: FlowInfo) -> list[Finding]:
        findings = analyze_flow(flow, self.catalog)
        filtered = [f for f in findings if self._is_rule_applicable(f.rule, flow.name, flow.api_version)]
        return _sorted(filtered)

    def analyze_object(self, obj: ObjectInfo) -> list[Finding]:
        findings = analyze_object(obj, self.catalog)
        filtered = [f for f in findings if self._is_rule_applicable(f.rule, obj.api_name, obj.api_version)]
        return _sorted(filtered)

    def analyze_validation_rule(
        self, vr: ValidationRuleInfo, object_name: str
    ) -> list[Finding]:
        findings = analyze_validation_rule(vr, object_name, self.catalog)
        # Pour les VR, on peut exclure soit par "Objet.NomVR", soit juste "NomVR"
        vr_full_name = f"{object_name}.{vr.full_name}"
        filtered = [
            f for f in findings 
            if self._is_rule_applicable(f.rule, vr_full_name, vr.api_version)
            and self._is_rule_applicable(f.rule, vr.full_name, vr.api_version)
        ]
        return _sorted(filtered)

    def analyze_data_transform(
        self, name: str, xml_content: str
    ) -> list[Finding]:
        findings = analyze_data_transform(name, xml_content, self.catalog)
        # On n'a pas forcément la version d'API pour les Data Transforms ici
        filtered = [f for f in findings if self._is_rule_applicable(f.rule, name)]
        return _sorted(filtered)

    def analyze_agent(self, agent: AgentInfo) -> list[Finding]:
        rules = self.catalog.for_scope("agent")
        findings: list[Finding] = []
        for rule in rules:
            if not self._is_rule_applicable(rule, agent.name):
                continue
            if rule.id == "AGENT-READ-001" and not agent.description:
                findings.append(
                    Finding(
                        rule=rule,
                        target_kind="Agent",
                        target_name=agent.name,
                        message="L'agent ne dispose d'aucune description.",
                        source_path=agent.source_path,
                    )
                )
        return _sorted(findings)

    def analyze_prompt(self, prompt: GenAiPromptInfo) -> list[Finding]:
        rules = self.catalog.for_scope("prompt")
        findings: list[Finding] = []
        for rule in rules:
            if not self._is_rule_applicable(rule, prompt.name):
                continue
            if rule.id == "PROMPT-READ-001" and not prompt.description:
                findings.append(
                    Finding(
                        rule=rule,
                        target_kind="GenAiPromptTemplate",
                        target_name=prompt.name,
                        message="Le prompt template ne dispose d'aucune description.",
                        source_path=prompt.source_path,
                    )
                )
        return _sorted(findings)

    # ------------------------------------------------------------------ snapshot-level API

    def analyze_snapshot(self, snapshot: MetadataSnapshot) -> "AnalyzerReport":
        apex_findings: dict[str, list[Finding]] = {}
        for artifact in snapshot.apex_artifacts:
            apex_findings[artifact.name] = self.analyze_apex(artifact)

        for name, extra in _detect_apex_call_cycles(
            snapshot.apex_artifacts, self.catalog
        ).items():
            apex_findings.setdefault(name, []).extend(extra)
            apex_findings[name] = _sorted(apex_findings[name])

        flow_findings: dict[str, list[Finding]] = {}
        for flow in snapshot.flows:
            flow_findings[flow.name] = self.analyze_flow(flow)

        object_findings: dict[str, list[Finding]] = {}
        validation_findings: dict[str, list[Finding]] = {}
        for obj in snapshot.objects:
            findings = self.analyze_object(obj)
            object_findings[obj.api_name] = findings
            for vr in obj.validation_rules:
                vr_key = f"{obj.api_name}.{vr.full_name}"
                validation_findings[vr_key] = self.analyze_validation_rule(vr, obj.api_name)

        omni_findings: dict[str, list[Finding]] = {}
        for row in snapshot.inventory.get("omnistudio", []):
            source = str(row.get("Source") or "")
            folder = str(row.get("Dossier") or "").lower()
            file_type = str(row.get("TypeFichier") or "").lower()
            is_dt = (
                "omnidatatransform" in folder
                or file_type.endswith(".rpt-meta.xml")
            )
            if not (is_dt and source):
                continue
            candidate = snapshot.source_dir / source
            if not candidate.exists() or not candidate.is_file():
                continue
            try:
                xml_text = candidate.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
            name = str(row.get("Nom") or candidate.stem)
            omni_findings[name] = self.analyze_data_transform(name, xml_text)

        agent_findings: dict[str, list[Finding]] = {}
        for agent in snapshot.agents:
            agent_findings[agent.name] = self.analyze_agent(agent)

        prompt_findings: dict[str, list[Finding]] = {}
        for prompt in snapshot.gen_ai_prompts:
            prompt_findings[prompt.name] = self.analyze_prompt(prompt)

        return AnalyzerReport(
            apex=apex_findings,
            flows=flow_findings,
            objects=object_findings,
            validation_rules=validation_findings,
            data_transforms=omni_findings,
            agents=agent_findings,
            prompts=prompt_findings,
            rules_used=self.catalog.enabled,
        )


class AnalyzerReport:
    """Agrege les findings par type d'artefact et fournit des helpers de synthese."""

    def __init__(
        self,
        apex: dict[str, list[Finding]] | None = None,
        flows: dict[str, list[Finding]] | None = None,
        objects: dict[str, list[Finding]] | None = None,
        validation_rules: dict[str, list[Finding]] | None = None,
        data_transforms: dict[str, list[Finding]] | None = None,
        agents: dict[str, list[Finding]] | None = None,
        prompts: dict[str, list[Finding]] | None = None,
        rules_used: list | None = None,
    ) -> None:
        self.apex = apex or {}
        self.flows = flows or {}
        self.objects = objects or {}
        self.validation_rules = validation_rules or {}
        self.data_transforms = data_transforms or {}
        self.agents = agents or {}
        self.prompts = prompts or {}
        self.rules_used = rules_used or []

    def all_findings(self) -> list[Finding]:
        collected: list[Finding] = []
        for group in (
            self.apex,
            self.flows,
            self.objects,
            self.validation_rules,
            self.data_transforms,
            self.agents,
            self.prompts,
        ):
            for findings in group.values():
                collected.extend(findings)
        return collected

    def severity_counts(self) -> dict[str, int]:
        counts = {"Critical": 0, "Major": 0, "Minor": 0, "Info": 0}
        for finding in self.all_findings():
            key = finding.rule.severity
            counts[key] = counts.get(key, 0) + 1
        return counts

    def rule_counts(self) -> dict[str, int]:
        counts: dict[str, int] = {}
        for finding in self.all_findings():
            counts[finding.rule.id] = counts.get(finding.rule.id, 0) + 1
        return counts

    def category_counts(self) -> dict[str, int]:
        counts = {"Trusted": 0, "Easy": 0, "Adaptable": 0}
        for finding in self.all_findings():
            counts[finding.rule.category] = counts.get(finding.rule.category, 0) + 1
        return counts


# ---------------------------------------------------------------------------- helpers


def _sorted(findings: list[Finding]) -> list[Finding]:
    return sorted(findings, key=lambda f: (SEVERITY_ORDER.get(f.rule.severity, 99), f.rule.id))


# ---------------------------------------------------------------------------- call-graph / cycles


def _detect_apex_call_cycles(
    artifacts: list[ApexArtifact], catalog: RuleCatalog
) -> dict[str, list[Finding]]:
    """Detecte les cycles d'appels entre classes Apex (APEX-REL-003).

    La detection construit un graphe d'appels (ClassName -> {ClassName appelee}) base sur
    les identifiants en PascalCase mentionnes dans le code (apres retrait des commentaires
    et chaines litterales) puis applique l'algorithme de Tarjan pour extraire les SCCs.
    Les composantes de taille >= 2, ou les auto-boucles, remontent comme findings.
    """
    rule = catalog.get("APEX-REL-003")
    if not rule or not rule.enabled:
        return {}

    classes = [a for a in artifacts if a.kind == "class"]
    class_names = {a.name for a in classes}
    if len(class_names) < 2:
        return {}

    graph: dict[str, set[str]] = {name: set() for name in class_names}
    for artifact in classes:
        stripped = _strip_comments_and_strings(artifact.body)
        mentioned = {m for m in IDENTIFIER_RE.findall(stripped) if m in class_names}
        mentioned.discard(artifact.name)
        graph[artifact.name] = mentioned

    cycles = _find_cycles(graph)
    if not cycles:
        return {}

    findings_by_class: dict[str, list[Finding]] = {}
    for cycle in cycles:
        cycle_sorted = sorted(cycle)
        cycle_label = " -> ".join(cycle_sorted + [cycle_sorted[0]])
        details = [
            f"Classes participant au cycle : {', '.join(cycle_sorted)}.",
            f"Chaine simplifiee : {cycle_label}.",
        ]
        for cls in cycle_sorted:
            artifact = next((a for a in classes if a.name == cls), None)
            others = [c for c in cycle_sorted if c != cls]
            message = (
                "Classe impliquee dans un cycle d'appels avec "
                + (", ".join(others) if others else "elle-meme")
                + "."
            )
            finding = Finding(
                rule=rule,
                target_kind="ApexClass",
                target_name=cls,
                message=message,
                details=list(details),
                source_path=artifact.source_path if artifact else None,
            )
            findings_by_class.setdefault(cls, []).append(finding)
    return findings_by_class


def _find_cycles(graph: dict[str, set[str]]) -> list[list[str]]:
    """Retourne les composantes fortement connexes >= 2 noeuds (ou auto-boucles) via Tarjan."""
    index_counter = [0]
    stack: list[str] = []
    on_stack: dict[str, bool] = {}
    index: dict[str, int] = {}
    lowlink: dict[str, int] = {}
    sccs: list[list[str]] = []

    def strongconnect(node: str) -> None:
        index[node] = index_counter[0]
        lowlink[node] = index_counter[0]
        index_counter[0] += 1
        stack.append(node)
        on_stack[node] = True

        for neighbour in graph.get(node, set()):
            if neighbour not in graph:
                continue
            if neighbour not in index:
                strongconnect(neighbour)
                lowlink[node] = min(lowlink[node], lowlink[neighbour])
            elif on_stack.get(neighbour):
                lowlink[node] = min(lowlink[node], index[neighbour])

        if lowlink[node] == index[node]:
            component: list[str] = []
            while True:
                w = stack.pop()
                on_stack[w] = False
                component.append(w)
                if w == node:
                    break
            if len(component) >= 2:
                sccs.append(component)
            elif node in graph.get(node, set()):
                sccs.append(component)

    for node in list(graph.keys()):
        if node not in index:
            strongconnect(node)
    return sccs
