from __future__ import annotations

import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from src.ui.application import Application


def show_exclusion_screen(app: Application) -> None:
    """Create and show the exclusion management window."""
    ExclusionScreen(app)


class ExclusionScreen:
    def __init__(self, app: Application) -> None:
        self.app = app
        self.window = tk.Toplevel(app)
        self.window.title(app._t("exclusions_title"))
        self.window.geometry("1000x700")
        app._configure_secondary_window(self.window)

        self.current_file = tk.StringVar(value=app.exclusion_file_var.get())
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", lambda *args: self._apply_filter())
        self.metadata_data: list[tuple] = []
        self.rules_data: list[tuple] = []
        self.sort_state: dict[str, tuple[str, bool]] = {
            "metadata": ("type", False), # (column, reverse)
            "rules": ("type", False)
        }

        self._build_ui()
        if self.current_file.get():
            self._load_data()

    def _build_ui(self) -> None:
        # Main container
        main_frame = ttk.Frame(self.window, padding=16)
        main_frame.pack(fill="both", expand=True)

        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill="x", pady=(0, 12))
        
        ttk.Label(
            header_frame,
            text=self.app._t("exclusions_title"),
            font=("Segoe UI", 14, "bold"),
        ).pack(anchor="w")
        
        ttk.Label(
            header_frame,
            text=self.app._t("exclusions_description"),
            wraplength=800,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        # File selection
        file_frame = ttk.LabelFrame(main_frame, text=self.app._t("exclusions_file_label"), padding=10)
        file_frame.pack(fill="x", pady=(0, 12))
        
        file_row = ttk.Frame(file_frame)
        file_row.pack(fill="x")
        
        ttk.Entry(file_row, textvariable=self.current_file).pack(side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(file_row, text=self.app._t("exclusions_browse"), command=self._browse_file).pack(side="left")

        # Tabs
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True)

        self.metadata_tab = ttk.Frame(self.notebook, padding=10)
        self.rules_tab = ttk.Frame(self.notebook, padding=10)
        
        self.notebook.add(self.metadata_tab, text=self.app._t("exclusions_tab_metadata"))
        self.notebook.add(self.rules_tab, text=self.app._t("exclusions_tab_rules"))

        self._build_metadata_tab()
        self._build_rules_tab()

        # Footer buttons
        footer_frame = ttk.Frame(main_frame, padding=(0, 12, 0, 0))
        footer_frame.pack(fill="x")
        
        ttk.Button(
            footer_frame,
            text=self.app._t("configuration_close"),
            command=self.window.destroy,
        ).pack(side="right")
        
        ttk.Button(
            footer_frame,
            text=self.app._t("exclusions_save"),
            command=self._save_data,
        ).pack(side="right", padx=(0, 8))

        # Add/Edit/Delete buttons in the footer
        ttk.Button(
            footer_frame,
            text=self.app._t("exclusions_delete"),
            command=self._on_delete,
        ).pack(side="left", padx=(0, 8))

        ttk.Button(
            footer_frame,
            text=self.app._t("configuration_ai_tags_edit"),
            command=self._on_edit,
        ).pack(side="left", padx=(0, 8))

        ttk.Button(
            footer_frame,
            text=self.app._t("exclusions_add"),
            command=self._on_add,
        ).pack(side="left")

    def _on_add(self) -> None:
        if self.notebook.index("current") == 0:
            self._add_metadata_row()
        else:
            self._add_rules_row()

    def _on_edit(self) -> None:
        if self.notebook.index("current") == 0:
            self._edit_metadata_row()
        else:
            self._edit_rules_row()

    def _on_delete(self) -> None:
        if self.notebook.index("current") == 0:
            self._delete_metadata_row()
        else:
            self._delete_rules_row()

    def _build_metadata_tab(self) -> None:
        # Filter row
        filter_frame = ttk.Frame(self.metadata_tab)
        filter_frame.pack(fill="x", pady=(0, 8))
        
        ttk.Label(filter_frame, text=self.app._t("exclusions_filter_label")).pack(side="left", padx=(0, 8))
        ttk.Entry(filter_frame, textvariable=self.filter_var).pack(side="left", fill="x", expand=True)

        # Table
        columns = ("type", "pattern", "comment")
        self.metadata_tree = ttk.Treeview(self.metadata_tab, columns=columns, show="headings")
        
        for col in columns:
            self.metadata_tree.heading(
                col, 
                text=self.app._t(f"exclusions_column_{col}"),
                command=lambda c=col: self._sort_column("metadata", c)
            )
        
        self.metadata_tree.column("type", width=150)
        self.metadata_tree.column("pattern", width=350)
        self.metadata_tree.column("comment", width=300)
        
        scrollbar = ttk.Scrollbar(self.metadata_tab, orient="vertical", command=self.metadata_tree.yview)
        self.metadata_tree.configure(yscrollcommand=scrollbar.set)
        
        self.metadata_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _build_rules_tab(self) -> None:
        # Filter row
        filter_frame = ttk.Frame(self.rules_tab)
        filter_frame.pack(fill="x", pady=(0, 8))
        
        ttk.Label(filter_frame, text=self.app._t("exclusions_filter_label")).pack(side="left", padx=(0, 8))
        self.rules_filter_var = tk.StringVar()
        self.rules_filter_var.trace_add("write", lambda *args: self._apply_rules_filter())
        ttk.Entry(filter_frame, textvariable=self.rules_filter_var).pack(side="left", fill="x", expand=True)

        # Table
        columns = ("type", "metadata", "rule", "comment")
        self.rules_tree = ttk.Treeview(self.rules_tab, columns=columns, show="headings")
        
        for col in columns:
            self.rules_tree.heading(
                col, 
                text=self.app._t(f"exclusions_column_{col}"),
                command=lambda c=col: self._sort_column("rules", c)
            )
        
        self.rules_tree.column("type", width=150)
        self.rules_tree.column("metadata", width=250)
        self.rules_tree.column("rule", width=200)
        self.rules_tree.column("comment", width=200)
        
        scrollbar = ttk.Scrollbar(self.rules_tab, orient="vertical", command=self.rules_tree.yview)
        self.rules_tree.configure(yscrollcommand=scrollbar.set)
        
        self.rules_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _browse_file(self) -> None:
        path = filedialog.askopenfilename(
            title=self.app._t("choose_exclusion_file"),
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.current_file.set(path)
            self._load_data()

    def _load_data(self) -> None:
        path = Path(self.current_file.get())
        if not path.exists():
            return

        try:
            self.metadata_tree.delete(*self.metadata_tree.get_children())
            self.rules_tree.delete(*self.rules_tree.get_children())
            
            wb = openpyxl.load_workbook(path, data_only=True)
            
            # Load metadata exclusions
            self.metadata_data.clear()
            sheet_metadata = None
            for name in wb.sheetnames:
                if name.strip().lower() == "hors analyse":
                    sheet_metadata = wb[name]
                    break
            
            if sheet_metadata:
                for row in sheet_metadata.iter_rows(min_row=1, values_only=True):
                    if not any(row): continue
                    if str(row[0]).startswith("#"): continue
                    # Ensure row has at least 3 elements (Type, Pattern, Comment)
                    full_row = list(row)
                    while len(full_row) < 3:
                        full_row.append("")
                    self.metadata_data.append(tuple(full_row))

            # Load rule exclusions
            self.rules_data.clear()
            sheet_rules = None
            for name in wb.sheetnames:
                if name.strip().lower() in ("exclusions regles", "exclusions règles"):
                    sheet_rules = wb[name]
                    break
            
            if sheet_rules:
                for row in sheet_rules.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    # Ensure row has at least 4 elements (Type, Metadata, Rule, Comment)
                    full_row = list(row)
                    while len(full_row) < 4:
                        full_row.append("")
                    self.rules_data.append(tuple(full_row))
            
            wb.close()
            self._refresh_trees()
        except Exception as e:
            messagebox.showerror(self.app._t("error_title"), f"{self.app._t('exclusions_load_error')}\n{e}")

    def _refresh_trees(self) -> None:
        self._apply_filter()
        self._apply_rules_filter()

    def _sort_column(self, tab_key: str, col: str) -> None:
        """Sort the in-memory data and refresh the tree."""
        current_col, reverse = self.sort_state[tab_key]
        if current_col == col:
            reverse = not reverse
        else:
            reverse = False
        
        self.sort_state[tab_key] = (col, reverse)
        
        # Update headings to show sort direction
        tree = self.metadata_tree if tab_key == "metadata" else self.rules_tree
        cols = ["type", "pattern", "comment"] if tab_key == "metadata" else ["type", "metadata", "rule", "comment"]
        
        for c in cols:
            text = self.app._t(f"exclusions_column_{c}")
            if c == col:
                text += " ↑" if not reverse else " ↓"
            tree.heading(c, text=text)

        if tab_key == "metadata":
            col_idx = cols.index(col)
            self.metadata_data.sort(key=lambda x: str(x[col_idx] or "").lower(), reverse=reverse)
            self._apply_filter()
        else:
            col_idx = cols.index(col)
            self.rules_data.sort(key=lambda x: str(x[col_idx] or "").lower(), reverse=reverse)
            self._apply_rules_filter()

    def _apply_filter(self) -> None:
        query = self.filter_var.get().lower()
        self.metadata_tree.delete(*self.metadata_tree.get_children())
        
        for row in self.metadata_data:
            match = False
            for cell in row:
                if query in str(cell or "").lower():
                    match = True
                    break
            if match or not query:
                self.metadata_tree.insert("", "end", values=row)

    def _apply_rules_filter(self) -> None:
        query = self.rules_filter_var.get().lower()
        self.rules_tree.delete(*self.rules_tree.get_children())
        
        for row in self.rules_data:
            match = False
            for cell in row:
                if query in str(cell or "").lower():
                    match = True
                    break
            if match or not query:
                self.rules_tree.insert("", "end", values=row)

    def _add_metadata_row(self) -> None:
        def on_save(values):
            self.metadata_data.append(values)
            self._apply_filter()
        self._edit_row_dialog(self.metadata_tree, ["type", "pattern", "comment"], on_save_callback=on_save)

    def _edit_metadata_row(self) -> None:
        selected = self.metadata_tree.selection()
        if not selected: return
        item = selected[0]
        old_values = self.metadata_tree.item(item)["values"]
        
        def on_save(new_values):
            # Find and replace in memory data
            for i, row in enumerate(self.metadata_data):
                if list(row) == list(old_values):
                    self.metadata_data[i] = new_values
                    break
            self._apply_filter()
            
        self._edit_row_dialog(self.metadata_tree, ["type", "pattern", "comment"], old_values, item, on_save)

    def _add_rules_row(self) -> None:
        def on_save(values):
            self.rules_data.append(values)
            self._apply_rules_filter()
        self._edit_row_dialog(self.rules_tree, ["type", "metadata", "rule", "comment"], on_save_callback=on_save)

    def _edit_rules_row(self) -> None:
        selected = self.rules_tree.selection()
        if not selected: return
        item = selected[0]
        old_values = self.rules_tree.item(item)["values"]
        
        def on_save(new_values):
            # Find and replace in memory data
            for i, row in enumerate(self.rules_data):
                if list(row) == list(old_values):
                    self.rules_data[i] = new_values
                    break
            self._apply_rules_filter()
            
        self._edit_row_dialog(self.rules_tree, ["type", "metadata", "rule", "comment"], old_values, item, on_save)

    def _delete_metadata_row(self) -> None:
        selected = self.metadata_tree.selection()
        if not selected: return
        if messagebox.askyesno(self.app._t("info_title"), self.app._t("exclusions_confirm_delete")):
            for item in selected:
                values = self.metadata_tree.item(item)["values"]
                # Remove from memory data
                self.metadata_data = [r for row in self.metadata_data if list(row) != list(values)]
                self.metadata_tree.delete(item)

    def _delete_rules_row(self) -> None:
        selected = self.rules_tree.selection()
        if not selected: return
        if messagebox.askyesno(self.app._t("info_title"), self.app._t("exclusions_confirm_delete")):
            for item in selected:
                values = self.rules_tree.item(item)["values"]
                # Remove from memory data
                self.rules_data = [r for row in self.rules_data if list(row) != list(values)]
                self.rules_tree.delete(item)

    def _edit_row_dialog(self, tree: ttk.Treeview, fields: list[str], initial_values: list = None, item_id: str = None, on_save_callback: callable = None) -> None:
        dialog = tk.Toplevel(self.window)
        dialog.title(self.app._t("exclusions_add") if item_id is None else self.app._t("configuration_ai_tags_edit"))
        dialog.geometry("400x350")
        self.app._configure_secondary_window(dialog)
        
        vars = {}
        for i, field in enumerate(fields):
            ttk.Label(dialog, text=self.app._t(f"exclusions_column_{field}")).grid(row=i, column=0, padx=10, pady=10, sticky="w")
            
            initial_val = initial_values[i] if initial_values and i < len(initial_values) else ""
            
            if field == "type":
                from src.parsers.salesforce_parser import SalesforceMetadataParser
                categories = sorted(set(SalesforceMetadataParser.CATEGORY_ALIASES.values()))
                var = tk.StringVar(value=initial_val)
                combo = ttk.Combobox(dialog, textvariable=var, values=categories, state="readonly", width=30)
                combo.grid(row=i, column=1, padx=10, pady=10, sticky="ew")
                vars[field] = var
            else:
                var = tk.StringVar(value=initial_val)
                ttk.Entry(dialog, textvariable=var).grid(row=i, column=1, padx=10, pady=10, sticky="ew")
                vars[field] = var
        
        def save():
            values = [vars[f].get() for f in fields]
            if on_save_callback:
                on_save_callback(values)
            else:
                if item_id:
                    tree.item(item_id, values=values)
                else:
                    tree.insert("", "end", values=values)
            dialog.destroy()
            
        ttk.Button(dialog, text=self.app._t("configuration_save"), command=save).grid(row=len(fields), column=1, pady=20)
        dialog.columnconfigure(1, weight=1)

    def _save_data(self) -> None:
        path = self.current_file.get()
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            # Metadata sheet
            ws_meta = wb.active
            ws_meta.title = "Hors analyse"
            for row in self.metadata_data:
                ws_meta.append(row)

            # Rules sheet
            ws_rules = wb.create_sheet("Exclusions regles")
            ws_rules.append(["Type", "Nom Metadata", "ID Regle", "Commentaire"]) # Header
            for row in self.rules_data:
                ws_rules.append(row)

            wb.save(path)
            wb.close()
            messagebox.showinfo(self.app._t("info_title"), self.app._t("exclusions_saved"))
        except Exception as e:
            messagebox.showerror(self.app._t("error_title"), f"{self.app._t('exclusions_save_error')}\n{e}")
